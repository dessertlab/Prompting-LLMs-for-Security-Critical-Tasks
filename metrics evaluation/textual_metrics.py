import re
import evaluate
import pylcs
import openpyxl
from openpyxl.styles import Font
import nltk
import bleu_score
from rouge import Rouge
from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction

#FUNZIONI PER IL CALCOLO DELLE METRICHE
def edit_dist(hyp, ref):
    tmp = pylcs.edit_distance(hyp, ref)
    res_norm = 1 - (tmp / max(len(hyp), len(ref)))
    return res_norm


def calc_ed(hyps, refs):
    res_ed = sum([edit_dist(h, r) for h, r in zip(hyps, refs)]) / len(hyps)
    return "{0:.3f}".format(res_ed)


meteor_model = evaluate.load("meteor")
def calc_meteor(hyps, refs, meteor_model):
    res_meteor = meteor_model.compute(predictions=hyps, references=refs)
    return "{0:.3f}".format(res_meteor["meteor"])


def calc_rouge(hyps, refs, type):
    metrics = [type]
    rouge = Rouge(metrics=metrics)
    scores = rouge.get_scores(hyps, refs, avg=True)

    return round(scores[type]["f"], 3)


def calc_bleu(hyps, refs, i):

    bleu_tup = bleu_score.compute_bleu(
            [[x] for x in refs], hyps, smooth=False, max_order=i
    )
    bleu = bleu_tup[0]
    return "{0:.3f}".format(bleu)




#FUNZIONI DA APPLICARE SULLE CELLE EXCEL
def rimuovi_commenti(codice):
    if not codice:
        return ''
    # Rimuovi i commenti multi-riga
    codice_pulito = re.sub(r'"""(.|\n)*?"""', '', codice)
    # Rimuovi i commenti su singola riga
    codice_pulito = re.sub(r'#.*', '', codice_pulito)
    # Rimuovi le linee vuote
    codice_pulito = '\n'.join([linea for linea in codice_pulito.split('\n') if linea.strip()])
    return codice_pulito

def format_function_code(stringa_multipla):
    # Dividi la stringa in righe usando il carattere di newline come delimitatore
    righe = stringa_multipla.split("\n")

    # Unisci le righe in una sola stringa usando il carattere di newline come separatore
    stringa_unificata = "\\n".join(righe)

    return stringa_unificata

def estrai_corpo_funzione(codice):
    if not codice:
        return ''
    righe = codice.strip().split('\n')
    corpo = righe[1:] if len(righe) > 1 else []  # esclude la riga con def
    return '\n'.join([r for r in corpo if r.strip()])  # rimuove righe vuote




for i in range (1,5):
    if i == 1:
        file_name = 'VP Dataset GPT-4o.xlsx'
    elif i == 2: 
        file_name = 'VP Dataset Llama-3_1.xlsx'
    elif i == 3: 
        file_name = 'VP Dataset Phi-3_5.xlsx'   
    elif i == 4: 
        file_name = 'VP Dataset_Qwen-2_5.xlsx'  
        
                  
    print (f"\nMetriche di {file_name}\n")
    
    # Apro il file Excel    
    wb = openpyxl.load_workbook(file_name)
    sheet = wb['VP Functions']

    # Ottieni l'indice della colonna 'Ground Truth Code'
    col_indice_ground_truth_code = None
    for col in sheet.iter_cols():
        if col[0].value == 'Ground truth code':
            col_indice_ground_truth_code = col[0].column
            break
    # Verifica che la colonna 'Ground Truth Code' sia stata trovata
    if col_indice_ground_truth_code is None:
        print("Colonna 'Ground Truth Code' non trovata.")
        exit()

    # Trova indice della colonna "Dataset" nel foglio principale
    intestazioni = [cell.value for cell in sheet[1]]
    if "Dataset" not in intestazioni:
        print("Colonna 'Dataset' non trovata nel foglio 'VP Functions'.")
        exit()

    indice_col_dataset = intestazioni.index("Dataset") + 1

    metric_sheets = ['ED', 'METEOR'] + [f'BLEU-{i}' for i in range(1, 5)] + ["rouge-1", "rouge-2", "rouge-3", "rouge-4", "rouge-l"]

    colonne = [
        "Dataset", "Ground truth code", "Persona", "Template", "Question Refinement",
        "Alternative Approaches", "Context Manager", "Flipped Interaction (3)", "Flipped Interaction (4)",
        "Flipped Interaction (5)", "Iterative Prompting (3)", "Iterative Prompting (4)",
        "Iterative Prompting (5)", "Few Shots Prompting", "CoT Prompting", "Fact Check List",
        "Not interactive mix", "Interactive mix", "Baseline"
    ]

    # Crea i fogli se non esistono e aggiungi intestazione
    for sheet_name in metric_sheets:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        sheet_metric = wb[sheet_name]
        for idx, col_name in enumerate(colonne, start=1):
            sheet_metric.cell(row=1, column=idx, value=col_name)

    # Copia la colonna "Dataset" in tutti gli altri fogli di metriche

        foglio_metrica = wb[sheet_name]
        
        # Intestazione
        foglio_metrica.cell(row=1, column=1, value="Dataset")
        
        # Copia valori riga per riga (dalla seconda in poi)
        for row in range(2, sheet.max_row + 1):
            valore = sheet.cell(row=row, column=indice_col_dataset).value
            foglio_metrica.cell(row=row, column=1, value=valore)
    
    idx_baseline = colonne.index("Baseline") + 1

    # Iterazione sulle righe del foglio di lavoro
    for riga in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_indice_ground_truth_code, max_col=idx_baseline):
        print("Working on dataset row: ", riga[0].row)
        ground_truth_code = riga[0].value
        corpo_gt = estrai_corpo_funzione(ground_truth_code)
        v2 = format_function_code(rimuovi_commenti(corpo_gt))

        for i, cella in enumerate(riga[1:], start=1):
            corpo_cella = estrai_corpo_funzione(cella.value)
            v1 = rimuovi_commenti(corpo_cella)
            v3 = format_function_code(v1)

            if not ground_truth_code or not v1:
                continue

            hyps = [v3]
            refs = [v2]
            colonna_output = col_indice_ground_truth_code + i

            # ED
            risultato_ed = calc_ed(hyps, refs)
            wb['ED'].cell(row=cella.row, column=colonna_output, value=risultato_ed)

            # BLEU 1-4
            for b in range(1, 5):
                risultato_bleu = calc_bleu(hyps, refs, b)
                wb[f'BLEU-{b}'].cell(row=cella.row, column=colonna_output, value=risultato_bleu)

            # METEOR
            risultato_meteor = calc_meteor(hyps, refs, meteor_model)
            wb['METEOR'].cell(row=cella.row, column=colonna_output, value=risultato_meteor)

            # ROUGE
            for tipo in ["rouge-1", "rouge-2", "rouge-3", "rouge-4", "rouge-l"]:
                risultato_rouge = calc_rouge(hyps, refs, tipo)
                wb[tipo].cell(row=cella.row, column=colonna_output, value=risultato_rouge)

    # Font settings
    font_header = Font(name='Cambria', size=11, bold=True)
    font_regular = Font(name='Calibri', size=11, bold=False)

    for sheet in wb.worksheets:
        max_row = sheet.max_row
        max_col = sheet.max_column
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = sheet.cell(row=r, column=c)
                cell.font = font_header if r == 1 else font_regular

    wb.save(file_name)
    print("Operazione completata con successo.")
