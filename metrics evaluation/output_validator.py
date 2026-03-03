import pandas as pd
import ast
import sys
import re
from pathlib import Path
from expected_data import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import unicodedata
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

#Percentuale di successo per ogni versione
def pass_column(excel_path, sheet_name):
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        print(f"Foglio '{sheet_name}' non trovato in {excel_path}")
        return

    ws = wb[sheet_name]

    header_row = [str(ws[f'{get_column_letter(col)}1'].value or "").strip().lower() for col in range(1, ws.max_column + 1)]
    try:
        inizio_col = header_row.index("ground truth code") + 1
    except ValueError:
        print("Colonna 'Ground truth code' non trovata.")
        return

    success_col_idx = None
    for col_idx in range(inizio_col, ws.max_column + 1):
        header = str(ws[f'{get_column_letter(col_idx)}1'].value or "").strip().lower()
        if header in ["success %", "success"]:
            success_col_idx = col_idx
            break

    if success_col_idx is None:
        print("Colonna 'success' non trovata.")
        return

    ultima_riga = None
    for row_idx in range(2, ws.max_row + 2):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        if all((v is None or str(v).strip() == "") for v in row_values):
            ultima_riga = row_idx - 1
            break
    if ultima_riga is None:
        ultima_riga = ws.max_row

    for col_idx in range(inizio_col, success_col_idx):
        col_letter = get_column_letter(col_idx)
        pass_count = 0
        total = 0

        for row in range(2, ultima_riga + 1):
            val = ws[f'{col_letter}{row}'].value
            if val is not None and str(val).strip() != "":
                total += 1
                if str(val).strip().upper() == 'PASS':
                    pass_count += 1

        percentuale = (pass_count / total) * 100 if total > 0 else 0
        ws[f'{col_letter}{ultima_riga + 1}'] = f'{percentuale:.2f}'

    wb.save(excel_path)
    print(f"Percentuali aggiornate in '{sheet_name}'")

#Estrae la prima esecuzione da output
def flatten_first_output(obj):
    while isinstance(obj, (list, tuple)) and len(obj) == 1:
        obj = obj[0]
    return obj

def extract_keywords_from_any_object(obj, keywords):
    try:
        string_repr = str(obj)
        found = [kw for kw in keywords if kw.lower() in string_repr.lower()]
        return {
            "string": string_repr,
            "found_keywords": found,
            "error": False
        }
    except Exception as e:
        return {
            "string": f"[ERROR] {str(e)}",
            "found_keywords": [],
            "error": True
        }


def output_validation(folder_path, i):
    df = pd.read_excel(folder_path)
    results = []

    script_name = Path(folder_path).stem
    keywords = VALID_BY_SCRIPT.get(script_name, [])

    for idx, row in df.iterrows():
        if script_name in ["train17", "train22", "train26"]:
            results.append("")
            continue
        
        if idx == 0:
            results.append("PASS")
            continue

        output_raw = row["Outputs"]
        output_str = str(output_raw).strip()

        if output_str.startswith("[") and output_str.endswith("]"):
            output_str = output_str[1:-1].strip()

        #Divide in esecuzioni tramite il separatore ';'
        executions = [e.strip() for e in output_str.split(");") if e.strip()]
        output = executions[0] if executions else ""

        try:
            parsed = ast.literal_eval(output)
            if isinstance(parsed, (list, tuple)) and parsed:
                output = flatten_first_output(parsed)
            else:
                output = parsed
        except Exception:
            pass

        parsed_result = extract_keywords_from_any_object(output, keywords)      
        
        if parsed_result["error"] or "[error]" in parsed_result["string"].lower():
            result = "FAIL"
        elif script_name in ["train17"]:
            result = "FAIL"
        elif script_name == "train47":
            if (idx == 9 and int(i) == 2):
                result = "PASS"
            else:
                result = "FAIL"
        elif script_name in ["dev7", "train44","train46", "train47", "train50", "train53", "train58"]:
            if any(kw.lower() in parsed_result["string"].lower() for kw in keywords):
                result = "PASS"
            else:
                result = "FAIL"
        elif script_name == "test7":
            numbers = re.findall(r"\b\d+(?:\.\d+)?\b", parsed_result["string"])
            if len(numbers) == 1:
                result = "PASS"
            else:
                result = "FAIL"
        elif script_name == "train52" and int(i) == 2:
                result = "PASS"
        else:
            if all(kw.lower() in parsed_result["string"].lower() for kw in keywords):
                result = "PASS"
            else:
                result = "FAIL"

        results.append(result)

    # Calcolo percentuale di "PASS"
    valid_results = results[1:]
    total_valid = len(valid_results)
    num_pass = valid_results.count("PASS")
    success = round((num_pass / total_valid) * 100, 2) if total_valid > 0 else 0.0

    return results, success


def robustness_test(folder_path, valid_script, i):
    df = pd.read_excel(folder_path)
    results = []

    script_name = Path(folder_path).stem   
    edge = EDGE_BY_SCRIPT.get(script_name, [])
    invalid = INVALID_BY_SCRIPT.get(script_name, [])

    for idx, row in df.iterrows():
        if script_name in ["test6", "train14", "train15"]:
            results.append("")
            continue
            
        if not valid_script[idx] or valid_script[idx] == "FAIL":
            results.append("")
            continue
            
        else:
            output_raw = row["Outputs"]
            output_str = str(output_raw).strip()

            if output_str.startswith("[") and output_str.endswith("]"):
                output_str = output_str[1:-1].strip()

            executions = [e.strip() for e in output_str.split("); (")]
            executions = [e if (e.startswith("(") and e.endswith(")")) else f"({e.strip('()')})" for e in executions]
            
            #Casi in cui è stato fatto solo test con input valido e invalido senza edge
            if len(executions) == 2:
                executions.append(executions[1])
                executions[1] = executions[0]
            
            check = [] 
            
            for j in range(1, 3):  
                          
                keywords = edge if j == 1 else invalid
                
                try:
                    parsed = ast.literal_eval(executions[j])

                    if isinstance(parsed, (list, tuple)) and parsed:
                        output = flatten_first_output(parsed)
                    else:
                        output = parsed
                except Exception as e:
                    output = executions[j]
                
                parsed_result = extract_keywords_from_any_object(output, keywords)    

                if script_name == "train39":
                    if all(kw.lower() in parsed_result["string"].lower() for kw in keywords):
                        result = "PASS"
                    else:
                        result = "FAIL"
                elif j == 1 and (not parsed_result["found_keywords"]):
                    check.append("PASS")
                elif j == 2 and (not parsed_result["found_keywords"]):
                    check.append("FAIL")
                elif parsed_result["found_keywords"]:
                    check.append("PASS")
                
                else:
                    check.append("FAIL")

            if not check:
                results.append("")
            elif check == ["PASS", "PASS"]:
                results.append("PASS")
            else:
                results.append("FAIL")
        
    # Calcolo percentuale di "PASS" solo sulle versioni corrette
    valid_results = [r for r in results[:] if r in ("PASS", "FAIL")]
    total_valid = len(valid_results)
    num_pass = valid_results.count("PASS")
    success = round((num_pass / total_valid) * 100, 2) if total_valid > 0 else 0.0

    return results, success


def excels_maker(folder_path, output_file, i):
    if i == 3:
        colonne = [
            "Dataset", "Ground truth code", "Persona", "Template", "Question Refinement",
            "Alternative Approaches", "Context Manager", "Iterative Prompting (3)", "Iterative Prompting (4)",
            "Iterative Prompting (5)", "Few Shots Prompting", "CoT Prompting", "Fact Check List",
            "Not interactive mix", "Baseline", "Success %"
        ]
    else:
        colonne = [
            "Dataset", "Ground truth code", "Persona", "Template", "Question Refinement",
            "Alternative Approaches", "Context Manager", "Flipped Interaction (3)", "Flipped Interaction (4)",
            "Flipped Interaction (5)", "Iterative Prompting (3)", "Iterative Prompting (4)",
            "Iterative Prompting (5)", "Few Shots Prompting", "CoT Prompting", "Fact Check List",
            "Not interactive mix", "Interactive mix", "Baseline", "Success %"
        ]
    
    df_vp = pd.DataFrame(columns=colonne)
    df_rb = pd.DataFrame(columns=colonne)
    folder = Path(folder_path)

    for file in sorted(folder.glob("*.xlsx"), key=lambda f: f.name.lower()):
        results_vp, success_vp = output_validation(file, i)
        results_rb, success_rb = robustness_test(file, results_vp, i) 
        
        dataset_name = file.name.replace(".xlsx", "")     
        new_row_vp = {"Dataset": dataset_name, "Success %": success_vp}
        new_row_rb = {"Dataset": dataset_name, "Success %": success_rb}
        
        for i, col in enumerate(colonne[1:-1]):
            new_row_vp[col] = results_vp[i] if i < len(results_vp) else ""
            new_row_rb[col] = results_rb[i] if i < len(results_rb) else ""

        df_vp = pd.concat([df_vp, pd.DataFrame([new_row_vp])], ignore_index=True)
        df_rb = pd.concat([df_rb, pd.DataFrame([new_row_rb])], ignore_index=True)
 

    with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df_vp.to_excel(writer, sheet_name="Validation", index=False)
        df_rb.to_excel(writer, sheet_name="Robustness", index=False)
        
        
    pass_column(output_file, "Validation")
    pass_column(output_file, "Robustness")
    print(f"Salvato in: {output_file}\n")   


for i in range (4,5):
    if i == 1:
        folder_path = "./../docker environment/output_Llama-3_1"
        output_file= "VP Dataset Llama-3_1.xlsx"
    elif i == 2:
        folder_path = "./../docker environment/output_GPT-4o"
        output_file= "VP Dataset GPT-4o.xlsx"
    elif i == 3:
        folder_path = "./../docker environment/output_Phi-3_5"
        output_file= "VP Dataset Phi-3_5.xlsx"
    elif i == 4:
        folder_path = "./../docker environment/output_Qwen-2_5"
        output_file= "VP Dataset_Qwen-2_5.xlsx"
    
    excels_maker(folder_path, output_file, i)   
