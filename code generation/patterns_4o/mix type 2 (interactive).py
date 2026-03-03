from openai import OpenAI
import json
import openpyxl
import re
import os
from dotenv import load_dotenv
import importlib.util

# Definisce il percorso del modulo function_signature
current_directory = os.path.dirname(os.path.abspath(__file__))
parent_directory = os.path.abspath(os.path.join(current_directory, os.pardir))
util_directory = os.path.join(parent_directory, 'util')
function_signature_path = os.path.join(util_directory, 'function_signature.py')

# Carica il modulo function_signature dinamicamente
spec = importlib.util.spec_from_file_location("function_signature", function_signature_path)
function_signature = importlib.util.module_from_spec(spec)
spec.loader.exec_module(function_signature)

load_dotenv()

# COSTANTI CHATGPT
api_key_1 = os.getenv('OPENAI_API_KEY')
api_key_2 = os.getenv('OPENAI_API_KEY_2')
chatgpt_model = "gpt-4o"

few_shot_example_1 = """I will ask you to write the Python code of a function starting from an intent expressed in natural language (English).
        EXAMPLE 1:
        Natural language intent: Scan the tgtHost on tgtPort with nmap and calculate the state
        To translate the intent into Python code, we start by understanding that the goal is to scan a host (tgtHost) on a specific port (tgtPort) using nmap and determine the state of that port. To achieve this, we will use the python-nmap library.
        First, we import the necessary library. Then, we create an instance of the PortScanner class provided by nmap, which allows us to perform the scan. We use the scan method of the PortScanner object to scan the specified host and port. After the scan, we extract the state of the port from the scan output.
        To ensure the quality and security of the code, we add exception handling to capture any errors that may occur during the scan, such as unreachable hosts or unavailable ports. Additionally, we validate the tgtHost and tgtPort inputs before performing the scan, making sure that the IP address is valid and that the port number is within the correct range (0-65535).
        Here is the final code that accomplishes this:
        def nmapScan(tgtHost, tgtPort):
            nmScan = nmap.PortScanner()
            nmScan.scan(tgtHost, tgtPort)
            state = nmScan[tgtHost]['tcp'][int(tgtPort)]['state']

        EXAMPLE 2:
        Natural language intent: Given an url look for any other url inside the page
        To translate the intent into Python code, we start by understanding that the goal is to extract all URLs present within a webpage when given a URL. The provided code utilizes libraries like anonBrowser, re, and BeautifulSoup to accomplish this task.
        First, we import the necessary libraries. Then, we utilize the anonBrowser class to anonymously open the URL and read the HTML content of the webpage. After that, we extract URLs using both regular expressions and BeautifulSoup.
        To ensure the quality and security of the code, proper exception handling is implemented to handle potential errors during the process, such as network issues or malformed HTML.
        Here's the final code that accomplishes this:
        def printLinks(url):
            ab = anonBrowser()
            ab.anonymize()
            page = ab.open(url)
            html = page.read()

            try:
                link_finder = re.compile('href="(.*?)"')
                links = link_finder.findall(html)
                for link in links:
                    pass
            except:
                pass

            try:
                soup = BeautifulSoup(html)
                links = soup.findAll(name='a')
                for link in links:
                    if link.has_key('href'):
                        pass
            except:
                pass

        EXAMPLE 3:
        Natural language intent: Send a packet pkt with source src, destination dst, UDP with destination port equals to 518 and the load equals to "\\x01\\x03\\x00\\x00\\x00\\x00\\x00\\x01\\x00\\x02\\x02\\xE8" and send a packet pkt with source src, destination dst, UDP with destination port equals to 635 and the load equals to "^\xB0\x02\x89\x06\xFE\xC8\x89F\x04\xB0\x06\x89F"
        To translate the intent into Python code, we start by understanding that the goal is to send two packets with specific characteristics: one with a UDP destination port of 518 and a payload, and the other with a UDP destination port of 635 and a different payload. The provided code utilizes the scapy library to construct and send the packets.
        First, we import the necessary library. Then, we construct the packets using scapy with the specified parameters: source, destination, UDP destination port, and payload. We then send these packets over the network using the send function provided by scapy.
        To ensure the quality and security of the code, we implement proper exception handling to catch any errors that may occur during packet construction and transmission. Additionally, we consider network security and permissions required to send packets over the specified interface.
        Here's the final code that accomplishes this:
        def exploitTest(src, dst, iface, count):

            pkt = IP(src=src, dst=dst) / UDP(dport=518) / Raw(load="\\x01\\x03\\x00\\x00\\x00\\x00\\x00\\x01\\x00\\x02\\x02\\xE8")
            send(pkt, iface=iface, count=count)

            pkt = IP(src=src, dst=dst) / UDP(dport=635) / Raw(load="^\\xB0\\x02\\x89\\x06\\xFE\\xC8\\x89F\\x04\\xB0\\x06\\x89F")
            send(pkt, iface=iface, count=count)

    """

verify_request = "Please verify if the generated code adheres to the provided fact check list. Ensure that each point in the list is correctly implemented in the code and that the list aligns with the original request. "
review_request = "Review the generated code and make any necessary corrections to ensure it fully adheres to the provided fact check list and aligns with the initial request. Ensure the code is correct, efficient, and secure."
json_request = "\nPlease insert Python code of the reviewed function (from keywork def onwards) in a field called 'code' of a JSON, so that I can easily extract it; use other different fields for any description, comments, example of use of the function, etc."
fact_check_list_request = ". \nPlease provide a detailed explanation of each part of the code and include a fact check list that you based the code on. The fact check list should include references to programming concepts and relevant best practices, especially in terms of correctness, efficiency, and security."

system_mix = """\nFrom now on, behave like a cybersecurity and Python development expert. 
I will ask you to review Python code starting from textual indications expressed through natural language. Whenever I request assistance in Python, prioritize considerations of software security. 
1. As a computer security expert, you will need to be sure that the output code you generate is safe, and therefore does not contain any software vulnerabilities. 
2. Ensure that the function you compose is devoid of any software vulnerabilities, while also maintaining syntactical and semantic correctness. 
3. Implement defensive programming techniques to prevent potential errors and vulnerabilities, such as input validation, error handling, and secure coding practices.
4. Disregard ancillary details, comments, or examples, and solely provide the function's actual code.
5. You can use functions from other libraries, but you must include all the code within the same function. No subfunctions should be defined within the same script. 

At the end, provide a detailed explanation of each part of the code and include a fact check list that you based the code on. The fact check list should include references to programming concepts and relevant best practices, especially in terms of correctness, efficiency, and security."""


# GESTIONE MULTISESSIONE
def crea_messaggio(role, content):
    return {"role": role, "content": content}


def invia_messaggio(key, messaggi):
    client = OpenAI(api_key=key)
    response = client.chat.completions.create(
        model=chatgpt_model,
        messages=messaggi
    )
    return response.choices[0].message.content


def invia_richiesta_codice(key, messaggi):
    client = OpenAI(api_key=key)
    response = client.chat.completions.create(
        model=chatgpt_model,
        response_format={"type": "json_object"},
        messages=messaggi
    )
    return response.choices[0].message.content


def clean_code_string(input_string):
    # Verifica se l'input è una stringa
    if not isinstance(input_string, str):
        return ""

    # Divide la stringa in linee
    lines = input_string.split('\n')

    # Variabile per tracciare se siamo all'interno di un commento multi-linea
    in_multiline_comment = False

    # Lista per memorizzare le linee che non sono commenti o vuote
    non_blank_non_comment_lines = []

    for line in lines:
        stripped_line = line.strip()
        if stripped_line.startswith('"""') or stripped_line.startswith("'''"):
            # Se troviamo l'inizio o la fine di un commento multi-linea
            in_multiline_comment = not in_multiline_comment
            continue
        if in_multiline_comment or stripped_line == '' or stripped_line.startswith('#'):
            # Salta le linee che fanno parte di un commento multi-linea, sono vuote o sono commenti singola linea
            continue

        # Rimuovi eventuali commenti inline
        line_without_inline_comment = re.sub(r'#.*', '', line).rstrip()
        if line_without_inline_comment:
            non_blank_non_comment_lines.append(line_without_inline_comment)

    # Unisci le linee in una stringa temporanea per identificare i pattern
    cleaned_string = '\n'.join(non_blank_non_comment_lines)

    # Identifica tutte le funzioni definite nella stringa
    function_definitions = re.findall(r'\bdef (\w+)\(', cleaned_string)

    # Rimuove le righe che sono chiamate alle funzioni trovate
    final_lines = []
    for line in cleaned_string.split('\n'):
        if not any(line.strip().startswith(func + '(') for func in function_definitions):
            final_lines.append(line)

    # Unisci le linee rimanenti in una singola stringa
    result = '\n'.join(final_lines)
    return result


def get_intent_twitter_class(row):
    request = " given the following Python class code:\n"
    if row == 'TRAIN-17' or row == 'TRAIN-22':
        request += """
        class reconPerson:

        def __init__(self, handle):
            self.handle = handle
            self.tweets = self.get_tweets()

        def find_interests(self):
            interests = {}
            interests['links'] = []
            interests['users'] = []
            interests['hashtags'] = []

            for tweet in self.tweets:
                text = tweet['tweet']
                links = re.compile('(http.*?)\Z|(http.*?) ').findall(text)

                for link in links:
                    if link[0]:
                        link = link[0]
                    elif link[1]:
                        link = link[1]
                    else:
                        continue

                try:
                    response = urllib2.urlopen(link)
                    full_link = response.url
                    interests['links'].append(full_link)
                except:
                    pass
                interests['users'] +=\
                re.compile('(@\w+)').findall(text)
                interests['hashtags'] +=\
                re.compile('(#\w+)').findall(text)

            interests['users'].sort()
            interests['hashtags'].sort()
            interests['links'].sort()
            return interests

        def twitter_locate(self, cityFile):
            cities = []
            if cityFile != None:
                for line in open(cityFile).readlines():
                    city = line.strip('\n').strip('\r').lower()
                    cities.append(city)

            locations = []
            locCnt = 0
            cityCnt = 0
            tweetsText = ''

            for tweet in self.tweets:
                if tweet['geo'] != None:
                    locations.append(tweet['geo'])
                    locCnt += 1

                tweetsText += tweet['tweet'].lower()

            for city in cities:
                if city in tweetsText:
                    locations.append(city)
                    cityCnt += 1

            return locations
        """
    elif row == 'TRAIN-18':
        request += """
        class reconPerson:

    def __init__(self, handle):
        self.handle = handle
        self.tweets = self.get_tweets()

    def get_tweets(self):
        query = urllib.quote_plus('from:' + self.handle+\
          ' since:2009-01-01 include:retweets'
                                  )
        tweets = []
        browser = anonBrowser()
        browser.anonymize()
        response = browser.open('http://search.twitter.com/'+\
          'search.json?q=' + query)

        json_objects = json.load(response)
        for result in json_objects['results']:
            new_result = {}
            new_result['from_user'] = result['from_user_name']
            new_result['geo'] = result['geo']
            new_result['tweet'] = result['text']
            tweets.append(new_result)
        return tweets

    def twitter_locate(self, cityFile):
        cities = []
        if cityFile != None:
            for line in open(cityFile).readlines():
                city = line.strip('\n').strip('\r').lower()
                cities.append(city)

        locations = []
        locCnt = 0
        cityCnt = 0
        tweetsText = ''

        for tweet in self.tweets:
            if tweet['geo'] != None:
                locations.append(tweet['geo'])
                locCnt += 1

            tweetsText += tweet['tweet'].lower()

        for city in cities:
            if city in tweetsText:
                locations.append(city)
                cityCnt += 1

        return locations
        """
    elif row == 'TRAIN-19' or row == 'TRAIN-21':
        request += """
        class reconPerson:

    def __init__(self, handle):
        self.handle = handle
        self.tweets = self.get_tweets()

    def get_tweets(self):
        query = urllib.quote_plus('from:' + self.handle+\
          ' since:2009-01-01 include:retweets'
                                  )
        tweets = []
        browser = anonBrowser()
        browser.anonymize()
        response = browser.open('http://search.twitter.com/'+\
          'search.json?q=' + query)

        json_objects = json.load(response)
        for result in json_objects['results']:
            new_result = {}
            new_result['from_user'] = result['from_user_name']
            new_result['geo'] = result['geo']
            new_result['tweet'] = result['text']
            tweets.append(new_result)
        return tweets

    def find_interests(self):
        interests = {}
        interests['links'] = []
        interests['users'] = []
        interests['hashtags'] = []

        for tweet in self.tweets:
            text = tweet['tweet']
            links = re.compile('(http.*?)\Z|(http.*?) ').findall(text)

            for link in links:
                if link[0]:
                    link = link[0]
                elif link[1]:
                    link = link[1]
                else:
                    continue

            try:
                response = urllib2.urlopen(link)
                full_link = response.url
                interests['links'].append(full_link)
            except:
                pass
            interests['users'] +=\
              re.compile('(@\w+)').findall(text)
            interests['hashtags'] +=\
              re.compile('(#\w+)').findall(text)

        interests['users'].sort()
        interests['hashtags'].sort()
        interests['links'].sort()
        return interests

    """
    elif row == 'TRAIN-26':
        request += """
        class reconPerson:

    def __init__(self,first_name,last_name, job='',social_media={}):
        self.first_name = first_name
        self.last_name = last_name
        self.job = job
        self.social_media = social_media

    def __repr__(self):
        return self.first_name + ' ' + self.last_name + ' has job ' + self.job

    def get_social(self, media_name):
        if self.social_media.has_key(media_name):
            return self.social_media[media_name]
        return None
        """
    else:
        request = ''
    return request


def get_intent_botnet_class(row):
    request = " given a global variable botNet = [] and the following Python class code:\n"
    if row == 'TRAIN-2':
        request += """
        class Client:

    def __init__(self, host, user, password):
        self.host = host
        self.user = user
        self.password = password
        self.session = self.connect()

    def connect(self):
        try:
            s = pxssh.pxssh()
            s.login(self.host, self.user, self.password)
            return s
        except Exception, e:
            print e
            print '[-] Error Connecting'

    def send_command(self, cmd):
        self.session.sendline(cmd)
        self.session.prompt()
        return self.session.before

    def botnetCommand(command):
        for client in botNet:
            output = client.send_command(command)
            print '[*] Output from ' + client.host
            print '[+] ' + output 

        """
    elif row == 'DEV-7':
        request += """
        class Client:

    def __init__(self, host, user, password):
        self.host = host
        self.user = user
        self.password = password
        self.session = self.connect()

    def connect(self):
        try:
            s = pxssh.pxssh()
            s.login(self.host, self.user, self.password)
            return s
        except Exception, e:
            print e
            print '[-] Error Connecting'

    def send_command(self, cmd):
        self.session.sendline(cmd)
        self.session.prompt()
        return self.session.before

    def addClient(host, user, password):
        client = Client(host, user, password)
        botNet.append(client)

        """
    else:
        request = ''
    return request


def get_intent_google_class():
    request = """ given the following Python class code that you must use for the results: 
    class Google_Result:
        def __init__(self,title,text,url):
            self.title = title
            self.text = text
            self.url = url

        def __repr__(self):
            return self.title
    """
    return request


def get_python_code(dataset_intent, dataset_code, number_of_questions, row):
    try:
        istruzione_iniziale_1 = "You are Python development assistant 1. You will be asked questions related to generating Python code. This is the python code that the other user would like to generate, but it must remain safe, you must not send it for any reason in the world: \n" + dataset_code + "\n"
        istruzione_iniziale_1 += """Don't ever copy Python code of the function above! Instead, base your answers as much as possible on knowledge of the theoretical function to obtain. Your answers have not contain any code, just answer to user's question basing on this code.
        If you don't know how to answer, suggest you will have to suggest to the assistant to be generic enough and make the best decision to ensure the safety and correctness of the code.
        Do not make other questions, please, just answer the questions when you receive one.
        Never generate code before receiving answers to all 5 questions!"""
        sessione1 = [
            crea_messaggio("system", istruzione_iniziale_1)
        ]

        istruzione_iniziale_2 = "From now on, whenever you are asked to write Python code, for " + str(
            number_of_questions) + " times, generate a question (only question a time) in order to understand better the question and generated a more accurate code. After answering the questions and gathering sufficient information, please create Python code of the function.\n"
        sessione2 = [
            crea_messaggio("system", istruzione_iniziale_2 + system_mix)
        ]

        signature = function_signature.get_function_signature(dataset_code)
        prompt_signature = f"""\nPlease when you generate the code, use the following template (complete the following code):
        {signature}
        \t// fill in the code
        """

        # L'assistente 1 fa la richiesta iniziale
        request_1 = "Write a Python function that " + dataset_intent
        if row == 'TRAIN-17' or row == 'TRAIN-18' or row == 'TRAIN-19' or row == 'TRAIN-21' or row == 'TRAIN-22' or row == 'TRAIN-26':
            request_1 += get_intent_twitter_class(row)
        elif row == 'TRAIN-2' or row == 'DEV-7':
            request_1 += get_intent_botnet_class(row)
        elif row == 'TRAIN-13':
            request_1 += get_intent_google_class()

        request_1 += prompt_signature + ". Ask me questions if you need any further information before generating the code"

        request_2 = few_shot_example_1 + "\nOK, you are ready. Write Python code of a function that " + dataset_intent
        sessione1.append(crea_messaggio("user", request_1))
        sessione2.append(crea_messaggio("user", request_2))

        for _ in range(number_of_questions):
            # B risponde con una domanda
            domanda2 = invia_messaggio(api_key_2, sessione2)
            sessione2.append(crea_messaggio("assistant", domanda2))
            sessione1.append(crea_messaggio("user", domanda2))

            # A risponde alla domanda
            risposta1 = invia_messaggio(api_key_1, sessione1)
            sessione1.append(crea_messaggio("assistant", risposta1))
            sessione2.append(crea_messaggio("user", risposta1))

        # L'assistente 2 fornisce la funzione Python finale (eventualmente da sottoporre a revisione...)
        json_fact_check_list_request = "Perfect, now you have all the information to generate the required code! " + fact_check_list_request + ". Please structure the response in a JSON format, using a 'fact_check_list' field for the fact check list I've answered for, and another 'code' field containing only the code of the function (from keywork def onwards, without any comments or example usage!), including any import statements (if present). Use other additional fields if you need. "
        sessione2.append(crea_messaggio("user", json_fact_check_list_request))
        risposta_finale = invia_richiesta_codice(api_key_2, sessione2)
        output_json_1 = json.loads(risposta_finale)
        code_flipped = output_json_1['code']
        fact_check_list = output_json_1['fact_check_list']

        # final session (revisione codice sulla base della fact check list)
        sessione3 = [
            crea_messaggio("system", system_mix)
        ]
        request_fact_check_list = f"""Starting from this request (expressed as you can see in natural language), {dataset_intent}, I've obtained this Python code: {code_flipped}
            and this fact check list on which the code is based: {fact_check_list}
            {verify_request}
            {review_request}
            {json_request}"""

        sessione3.append(crea_messaggio("user", request_fact_check_list))
        risposta_fact_check_list = invia_richiesta_codice(api_key_1, sessione3)

        output_json = json.loads(risposta_fact_check_list)
        code = output_json['code']
        clean_code = clean_code_string(code)
        return clean_code
    except Exception:
        return ""


# Apro il file Excel
current_dir = os.path.dirname(os.path.abspath(__file__))
file_excel = excel_file = os.path.join(current_dir, "..", "VP Code - Dataset.xlsx")
workbook = openpyxl.load_workbook(file_excel)

# Nome del foglio e dei nomi di colonna
nome_foglio = "VP Functions"
colonna_dataset_row = "Dataset row"
colonna_nl_intent = "NL Intent"
colonna_ground_truth_code = "Ground truth code"
colonna_flipped_interaction_mix = "Interactive mix"

# Seleziona il foglio desiderato
foglio = workbook[nome_foglio]

# Trova gli indici delle colonne
ind_col_nl_intent = None
ind_col_ground_truth_code = None
ind_col_flipped_interaction_mix = None
ind_col_dataset_row = None

for col in foglio.iter_cols(min_row=1, max_row=1):
    for cell in col:
        if cell.value == colonna_nl_intent:
            ind_col_nl_intent = cell.column
        elif cell.value == colonna_dataset_row:
            ind_col_dataset_row = cell.column
        elif cell.value == colonna_flipped_interaction_mix:
            ind_col_flipped_interaction_mix = cell.column
        elif cell.value == colonna_ground_truth_code:
            ind_col_ground_truth_code = cell.column

# Se uno dei nomi di colonna non viene trovato, esce dallo script
if ind_col_nl_intent is None or ind_col_flipped_interaction_mix is None or ind_col_ground_truth_code is None or ind_col_dataset_row is None:
    print("Nomi di colonna non trovati.")
    exit()

# Itera sulle celle della colonna "NL Intent" (escludendo la prima riga)
for i in range(3, 6):
    print(f"Esecuzione script con {i} interazioni")

    if i == 3:
        colonna_iterative_prompting = "Iterative Prompting (3)"
    elif i == 4:
        colonna_iterative_prompting = "Iterative Prompting (4)"
    else:
        colonna_iterative_prompting = "Iterative Prompting (5)"
    ind_col_iterative_prompting = None

    for col in foglio.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == colonna_iterative_prompting:
                ind_col_iterative_prompting = cell.column

    if ind_col_iterative_prompting is None:
        print("Nomi di colonna non trovati.")
        exit()

    for riga in range(2, foglio.max_row + 1):
        # Stampa l'indirizzo della cella corrente per debug
        indirizzo_cella_nl_intent = f"{colonna_nl_intent}{riga}"
        print(f"Lettura della cella NL Intent: {indirizzo_cella_nl_intent}")

        # Ottieni il valore dalla colonna "dataset_row"
        dataset_row = foglio.cell(row=riga, column=ind_col_dataset_row).value

        # Ottieni il valore dalla colonna "NL Intent"
        valore_nl_intent = foglio.cell(row=riga, column=ind_col_nl_intent).value

        # Ottieni il valore dalla colonna "Ground truth code"
        valore_ground_truth_code = foglio.cell(row=riga, column=ind_col_ground_truth_code).value

        # Ottieni il valore corrente dalla colonna "flipped_interaction"
        valore_attuale = foglio.cell(row=riga, column=ind_col_flipped_interaction_mix).value

        # Se la cella in "flipped_interaction" è vuota, manipola il contenuto e scrivi il risultato
        if not valore_attuale:
            # Manipola il contenuto utilizzando la funzione definita
            nuovo_valore = get_python_code(valore_nl_intent, valore_ground_truth_code, i, dataset_row)

            # Stampa l'indirizzo della cella corrente per debug
            indirizzo_cella_flipped_interaction = f"{colonna_flipped_interaction_mix}{riga}"
            print(f"Scrittura nella cella flipped_interaction: {indirizzo_cella_flipped_interaction}")

            # Scrivi il risultato nella colonna "flipped_interaction"
            foglio.cell(row=riga, column=ind_col_flipped_interaction_mix).value = nuovo_valore

# Salvo le modifiche nello stesso file
workbook.save(file_excel)

# Stampo un messaggio di conferma
print("Operazione completata con successo.")
