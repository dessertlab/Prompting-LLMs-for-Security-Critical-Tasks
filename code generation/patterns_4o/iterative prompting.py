import openpyxl
import re
import json
import os
from openai import OpenAI
from dotenv import load_dotenv
import importlib.util

load_dotenv()

# Definisce il percorso del modulo function_signature
current_directory = os.path.dirname(os.path.abspath(__file__))
parent_directory = os.path.abspath(os.path.join(current_directory, os.pardir))
util_directory = os.path.join(parent_directory, 'util')
function_signature_path = os.path.join(util_directory, 'function_signature.py')

# Carica il modulo function_signature dinamicamente
spec = importlib.util.spec_from_file_location("function_signature", function_signature_path)
function_signature = importlib.util.module_from_spec(spec)
spec.loader.exec_module(function_signature)


# Imposta la tua chiave API di OpenAI
openai_api_key = os.getenv('OPENAI_API_KEY')
client = OpenAI(api_key=openai_api_key)

json_request = ". \nPlease insert Python code of the function (from keywork def onwards) in a field called 'code' of a JSON, so that I can easily extract it; use other different fields for any description, comments, example of use of the function, etc."

def crea_messaggio(role, content):
    return {"role": role, "content": content}

def invia_richiesta_codice(client, messaggi):
    response = client.chat.completions.create(
        model="gpt-4o",
        response_format = { "type": "json_object" },
        messages=messaggi
    )
    return response.choices[0].message.content

def clean_code_string(input_string):
    # Verifica se l'input è una stringa
    if not isinstance(input_string, str):
        return ""

    # Divide la stringa in linee
    lines = input_string.split('\n')

    # Variabile per tracciare se siamo all'interno di un commento multi-linea
    in_multiline_comment = False

    # Lista per memorizzare le linee che non sono commenti o vuote
    non_blank_non_comment_lines = []

    for line in lines:
        stripped_line = line.strip()
        if stripped_line.startswith('"""') or stripped_line.startswith("'''"):
            # Se troviamo l'inizio o la fine di un commento multi-linea
            in_multiline_comment = not in_multiline_comment
            continue
        if in_multiline_comment or stripped_line == '' or stripped_line.startswith('#'):
            # Salta le linee che fanno parte di un commento multi-linea, sono vuote o sono commenti singola linea
            continue
        
        # Rimuovi eventuali commenti inline
        line_without_inline_comment = re.sub(r'#.*', '', line).rstrip()
        if line_without_inline_comment:
            non_blank_non_comment_lines.append(line_without_inline_comment)

    # Unisci le linee in una stringa temporanea per identificare i pattern
    cleaned_string = '\n'.join(non_blank_non_comment_lines)

    # Identifica tutte le funzioni definite nella stringa
    function_definitions = re.findall(r'\bdef (\w+)\(', cleaned_string)

    # Rimuove le righe che sono chiamate alle funzioni trovate
    final_lines = []
    for line in cleaned_string.split('\n'):
        if not any(line.strip().startswith(func + '(') for func in function_definitions):
            final_lines.append(line)

    # Unisci le linee rimanenti in una singola stringa
    result = '\n'.join(final_lines)
    return result


def get_intent_twitter_class(row):
    request = " given the following Python class code:\n"
    if row == 'TRAIN-17' or row == 'TRAIN-22':
        request += """
        class reconPerson:

        def __init__(self, handle):
            self.handle = handle
            self.tweets = self.get_tweets()

        def find_interests(self):
            interests = {}
            interests['links'] = []
            interests['users'] = []
            interests['hashtags'] = []

            for tweet in self.tweets:
                text = tweet['tweet']
                links = re.compile('(http.*?)\Z|(http.*?) ').findall(text)

                for link in links:
                    if link[0]:
                        link = link[0]
                    elif link[1]:
                        link = link[1]
                    else:
                        continue

                try:
                    response = urllib2.urlopen(link)
                    full_link = response.url
                    interests['links'].append(full_link)
                except:
                    pass
                interests['users'] +=\
                re.compile('(@\w+)').findall(text)
                interests['hashtags'] +=\
                re.compile('(#\w+)').findall(text)

            interests['users'].sort()
            interests['hashtags'].sort()
            interests['links'].sort()
            return interests

        def twitter_locate(self, cityFile):
            cities = []
            if cityFile != None:
                for line in open(cityFile).readlines():
                    city = line.strip('\n').strip('\r').lower()
                    cities.append(city)

            locations = []
            locCnt = 0
            cityCnt = 0
            tweetsText = ''

            for tweet in self.tweets:
                if tweet['geo'] != None:
                    locations.append(tweet['geo'])
                    locCnt += 1

                tweetsText += tweet['tweet'].lower()

            for city in cities:
                if city in tweetsText:
                    locations.append(city)
                    cityCnt += 1

            return locations
        """
    elif row == 'TRAIN-18':
        request += """
        class reconPerson:

    def __init__(self, handle):
        self.handle = handle
        self.tweets = self.get_tweets()

    def get_tweets(self):
        query = urllib.quote_plus('from:' + self.handle+\
          ' since:2009-01-01 include:retweets'
                                  )
        tweets = []
        browser = anonBrowser()
        browser.anonymize()
        response = browser.open('http://search.twitter.com/'+\
          'search.json?q=' + query)

        json_objects = json.load(response)
        for result in json_objects['results']:
            new_result = {}
            new_result['from_user'] = result['from_user_name']
            new_result['geo'] = result['geo']
            new_result['tweet'] = result['text']
            tweets.append(new_result)
        return tweets

    def twitter_locate(self, cityFile):
        cities = []
        if cityFile != None:
            for line in open(cityFile).readlines():
                city = line.strip('\n').strip('\r').lower()
                cities.append(city)

        locations = []
        locCnt = 0
        cityCnt = 0
        tweetsText = ''

        for tweet in self.tweets:
            if tweet['geo'] != None:
                locations.append(tweet['geo'])
                locCnt += 1

            tweetsText += tweet['tweet'].lower()

        for city in cities:
            if city in tweetsText:
                locations.append(city)
                cityCnt += 1

        return locations
        """
    elif row == 'TRAIN-19' or row == 'TRAIN-21':
        request += """
        class reconPerson:

    def __init__(self, handle):
        self.handle = handle
        self.tweets = self.get_tweets()

    def get_tweets(self):
        query = urllib.quote_plus('from:' + self.handle+\
          ' since:2009-01-01 include:retweets'
                                  )
        tweets = []
        browser = anonBrowser()
        browser.anonymize()
        response = browser.open('http://search.twitter.com/'+\
          'search.json?q=' + query)

        json_objects = json.load(response)
        for result in json_objects['results']:
            new_result = {}
            new_result['from_user'] = result['from_user_name']
            new_result['geo'] = result['geo']
            new_result['tweet'] = result['text']
            tweets.append(new_result)
        return tweets

    def find_interests(self):
        interests = {}
        interests['links'] = []
        interests['users'] = []
        interests['hashtags'] = []

        for tweet in self.tweets:
            text = tweet['tweet']
            links = re.compile('(http.*?)\Z|(http.*?) ').findall(text)

            for link in links:
                if link[0]:
                    link = link[0]
                elif link[1]:
                    link = link[1]
                else:
                    continue

            try:
                response = urllib2.urlopen(link)
                full_link = response.url
                interests['links'].append(full_link)
            except:
                pass
            interests['users'] +=\
              re.compile('(@\w+)').findall(text)
            interests['hashtags'] +=\
              re.compile('(#\w+)').findall(text)

        interests['users'].sort()
        interests['hashtags'].sort()
        interests['links'].sort()
        return interests

    """
    elif row == 'TRAIN-26':
        request += """
        class reconPerson:

    def __init__(self,first_name,last_name, job='',social_media={}):
        self.first_name = first_name
        self.last_name = last_name
        self.job = job
        self.social_media = social_media

    def __repr__(self):
        return self.first_name + ' ' + self.last_name + ' has job ' + self.job

    def get_social(self, media_name):
        if self.social_media.has_key(media_name):
            return self.social_media[media_name]
        return None
        """
    else:
        request = ''
    return request

def get_intent_botnet_class(row):
    request = " given a global variable botNet = [] and the following Python class code:\n"
    if row == 'TRAIN-2':
        request += """
        class Client:

    def __init__(self, host, user, password):
        self.host = host
        self.user = user
        self.password = password
        self.session = self.connect()

    def connect(self):
        try:
            s = pxssh.pxssh()
            s.login(self.host, self.user, self.password)
            return s
        except Exception, e:
            print e
            print '[-] Error Connecting'

    def send_command(self, cmd):
        self.session.sendline(cmd)
        self.session.prompt()
        return self.session.before

    def botnetCommand(command):
        for client in botNet:
            output = client.send_command(command)
            print '[*] Output from ' + client.host
            print '[+] ' + output 
        
        """
    elif row == 'DEV-7':
        request += """
        class Client:

    def __init__(self, host, user, password):
        self.host = host
        self.user = user
        self.password = password
        self.session = self.connect()

    def connect(self):
        try:
            s = pxssh.pxssh()
            s.login(self.host, self.user, self.password)
            return s
        except Exception, e:
            print e
            print '[-] Error Connecting'

    def send_command(self, cmd):
        self.session.sendline(cmd)
        self.session.prompt()
        return self.session.before

    def addClient(host, user, password):
        client = Client(host, user, password)
        botNet.append(client)

        """
    else:
        request = ''
    return request

def get_intent_google_class():
    request = """ given the following Python class code that you must use for the results: 
    class Google_Result:
        def __init__(self,title,text,url):
            self.title = title
            self.text = text
            self.url = url

        def __repr__(self):
            return self.title
    """
    return request
    


def get_python_code(input_text, codice_ground_truth_code, interazioni, row):
    try:
        sessione = [
            crea_messaggio("system", "You are a python assistant developer"),
        ]
        signature = function_signature.get_function_signature(codice_ground_truth_code)
        prompt_signature = f"""\nPlease when you generate the code, use the following template (complete the following code):
        {signature}
        \t// fill in the code
        \n
        """
        
        request = "Write Python code of a function that " + input_text 
        if row == 'TRAIN-17' or row == 'TRAIN-18' or row == 'TRAIN-19' or row == 'TRAIN-21' or row == 'TRAIN-22' or row == 'TRAIN-26':
            request += get_intent_twitter_class(row)
        elif row == 'TRAIN-2' or row == 'DEV-7':
            request += get_intent_botnet_class(row)
        elif row == 'TRAIN-13':
            request += get_intent_google_class()
                
        request += prompt_signature + json_request 
        
        sessione.append(crea_messaggio("user", request))
        risposta1 = invia_richiesta_codice(client, sessione)
        sessione.append(crea_messaggio("assistant", risposta1))
        code1 = json.loads(risposta1)['code']

        # interazione 1
        request_pep_8 = "Ensure the code adheres to PEP8, the official style guide for Python code (Python's best practices and coding conventions). This is the code to check: \n" + code1 + json_request
        sessione.append(crea_messaggio("user", request_pep_8))
        risposta2 = invia_richiesta_codice(client, sessione)
        sessione.append(crea_messaggio("assistant", risposta2))
        code2 = json.loads(risposta2)['code']

        # interazione 2
        request_defensive_programming = "Implement defensive programming techniques to prevent potential errors and vulnerabilities, such as input validation, error handling, and secure coding practices. This is the code to check: \n" + code2 + json_request
        sessione.append(crea_messaggio("user", request_defensive_programming))
        risposta3 = invia_richiesta_codice(client, sessione)
        sessione.append(crea_messaggio("assistant", risposta3))
        code3 = json.loads(risposta3)['code']

        # interazione 3
        request_updated_code = "Keep the code updated with the latest security patches and vulnerabilities for Python libraries and dependencies. This is the code to check: \n" + code3 + json_request
        sessione.append(crea_messaggio("user", request_updated_code))
        risposta4 = invia_richiesta_codice(client, sessione)
        sessione.append(crea_messaggio("assistant", risposta4))
        code4 = json.loads(risposta4)['code']
        
        
        if interazioni >= 4:
            # interazione 4
            request_review_code = "Subject the code to a rigorous code review process to identify potential security flaws, coding errors, and opportunities for improvement. This is the code to check: \n" + code4 + json_request
            sessione.append(crea_messaggio("user", request_review_code))
            risposta5 = invia_richiesta_codice(client, sessione)
            sessione.append(crea_messaggio("assistant", risposta5))
            code5 = json.loads(risposta5)['code']

            if interazioni >= 5:
                # interazione 5
                request_unnecessary_code = "Please remove any unnecessary comments, usage examples, classes, or other code from the function code. If not present, do not make any changes. This is the code to check: \n" + code5 + json_request
                sessione.append(crea_messaggio("user", request_unnecessary_code))
                risposta6 = invia_richiesta_codice(client, sessione)
                sessione.append(crea_messaggio("assistant", risposta6))
                final_code = json.loads(risposta6)['code']
            else:
                final_code = code5
        else:
            final_code = code4
        return clean_code_string(final_code)
    except Exception:
        return ""

# Apro il file Excel
current_dir = os.path.dirname(os.path.abspath(__file__))
file_excel = excel_file = os.path.join(current_dir, "..", "VP Code - Dataset.xlsx")
workbook = openpyxl.load_workbook(file_excel)

# Nome del foglio e dei nomi di colonna
nome_foglio = "VP Functions"
colonna_dataset_row = "Dataset row"
colonna_ground_truth_code = "Ground truth code"
colonna_nl_intent = "NL Intent"


# Seleziona il foglio desiderato
foglio = workbook[nome_foglio]

# Trova gli indici delle colonne
ind_col_ground_truth_code = None
ind_col_nl_intent = None
ind_col_dataset_row = None

for col in foglio.iter_cols(min_row=1, max_row=1):
    for cell in col:
        if cell.value == colonna_nl_intent:
            ind_col_nl_intent = cell.column
        elif cell.value == colonna_dataset_row:
            ind_col_dataset_row = cell.column
        elif cell.value == colonna_ground_truth_code:
            ind_col_ground_truth_code = cell.column

# Se uno dei nomi di colonna non viene trovato, esce dallo script
if ind_col_nl_intent is None or ind_col_ground_truth_code is None or ind_col_dataset_row is None:
    print("Nomi di colonna non trovati.")
    exit()

# Itera sulle celle della colonna "NL Intent" (escludendo la prima riga)

for i in range(3, 6):
    
    if i == 3:
        colonna_iterative_prompting = "Iterative Prompting (3)"
    elif i == 4:
        colonna_iterative_prompting = "Iterative Prompting (4)"
    else:
        colonna_iterative_prompting = "Iterative Prompting (5)"
    ind_col_iterative_prompting = None
    
    for col in foglio.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == colonna_iterative_prompting:
                ind_col_iterative_prompting = cell.column
                
    if ind_col_iterative_prompting is None:
        print("Nomi di colonna non trovati.")
        exit()
        
    for riga in range(2, foglio.max_row + 1):
        # Stampa l'indirizzo della cella corrente per debug
        indirizzo_cella_nl_intent = f"{colonna_nl_intent}{riga}"
        print(f"Lettura della cella NL Intent: {indirizzo_cella_nl_intent}")
        
        # Ottieni il valore dalla colonna "dataset_row"
        dataset_row = foglio.cell(row=riga, column=ind_col_dataset_row).value
    
        # Ottieni il valore dalla colonna "NL Intent"
        valore_nl_intent = foglio.cell(row=riga, column=ind_col_nl_intent).value
        
        # Ottieni il valore dalla colonna "Ground truth code"
        codice_ground_truth_code = foglio.cell(row=riga, column=ind_col_ground_truth_code).value

        # Ottieni il valore corrente dalla colonna "iterative_prompting"
        valore_attuale = foglio.cell(row=riga, column=ind_col_iterative_prompting).value
        
        # Se la cella in "iterative_prompting" è vuota, manipola il contenuto e scrivi il risultato
        if not valore_attuale:
            # Manipola il contenuto utilizzando la funzione definita
            nuovo_valore = get_python_code(valore_nl_intent, codice_ground_truth_code, i, dataset_row)

            # Stampa l'indirizzo della cella corrente per debug
            indirizzo_cella_iterative_prompting = f"{colonna_iterative_prompting}{riga}"
            print(f"Scrittura nella cella iterative_prompting: {indirizzo_cella_iterative_prompting}")
            
            # Scrivi il risultato nella colonna "iterative_prompting"
            foglio.cell(row=riga, column=ind_col_iterative_prompting).value = nuovo_valore


# Salvo le modifiche nello stesso file
workbook.save(file_excel)

# Stampo un messaggio di conferma
print("Operazione completata con successo.")
