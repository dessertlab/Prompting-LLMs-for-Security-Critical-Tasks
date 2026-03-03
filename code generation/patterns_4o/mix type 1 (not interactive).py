import openpyxl
import json
import re
from openai import OpenAI
import os
from dotenv import load_dotenv
import importlib.util

# Definisce il percorso del modulo function_signature
current_directory = os.path.dirname(os.path.abspath(__file__))
parent_directory = os.path.abspath(os.path.join(current_directory, os.pardir))
util_directory = os.path.join(parent_directory, 'util')
function_signature_path = os.path.join(util_directory, 'function_signature.py')

# Carica il modulo function_signature dinamicamente
spec = importlib.util.spec_from_file_location("function_signature", function_signature_path)
function_signature = importlib.util.module_from_spec(spec)
spec.loader.exec_module(function_signature)

load_dotenv()

# Imposta la tua chiave API di OpenAI
openai_api_key = os.getenv('OPENAI_API_KEY_2')
client = OpenAI(api_key=openai_api_key)

few_shot_example_1 = """I will ask you to write the Python code of a function starting from an intent expressed in natural language (English).
        EXAMPLE 1:
        Natural language intent: Scan the tgtHost on tgtPort with nmap and calculate the state
        To translate the intent into Python code, we start by understanding that the goal is to scan a host (tgtHost) on a specific port (tgtPort) using nmap and determine the state of that port. To achieve this, we will use the python-nmap library.
        First, we import the necessary library. Then, we create an instance of the PortScanner class provided by nmap, which allows us to perform the scan. We use the scan method of the PortScanner object to scan the specified host and port. After the scan, we extract the state of the port from the scan output.
        To ensure the quality and security of the code, we add exception handling to capture any errors that may occur during the scan, such as unreachable hosts or unavailable ports. Additionally, we validate the tgtHost and tgtPort inputs before performing the scan, making sure that the IP address is valid and that the port number is within the correct range (0-65535).
        Here is the final code that accomplishes this:
        def nmapScan(tgtHost, tgtPort):
            nmScan = nmap.PortScanner()
            nmScan.scan(tgtHost, tgtPort)
            state = nmScan[tgtHost]['tcp'][int(tgtPort)]['state']

        EXAMPLE 2:
        Natural language intent: Given an url look for any other url inside the page
        To translate the intent into Python code, we start by understanding that the goal is to extract all URLs present within a webpage when given a URL. The provided code utilizes libraries like anonBrowser, re, and BeautifulSoup to accomplish this task.
        First, we import the necessary libraries. Then, we utilize the anonBrowser class to anonymously open the URL and read the HTML content of the webpage. After that, we extract URLs using both regular expressions and BeautifulSoup.
        To ensure the quality and security of the code, proper exception handling is implemented to handle potential errors during the process, such as network issues or malformed HTML.
        Here's the final code that accomplishes this:
        def printLinks(url):
            ab = anonBrowser()
            ab.anonymize()
            page = ab.open(url)
            html = page.read()

            try:
                link_finder = re.compile('href="(.*?)"')
                links = link_finder.findall(html)
                for link in links:
                    pass
            except:
                pass

            try:
                soup = BeautifulSoup(html)
                links = soup.findAll(name='a')
                for link in links:
                    if link.has_key('href'):
                        pass
            except:
                pass

        EXAMPLE 3:
        Natural language intent: Send a packet pkt with source src, destination dst, UDP with destination port equals to 518 and the load equals to "\\x01\\x03\\x00\\x00\\x00\\x00\\x00\\x01\\x00\\x02\\x02\\xE8" and send a packet pkt with source src, destination dst, UDP with destination port equals to 635 and the load equals to "^\xB0\x02\x89\x06\xFE\xC8\x89F\x04\xB0\x06\x89F"
        To translate the intent into Python code, we start by understanding that the goal is to send two packets with specific characteristics: one with a UDP destination port of 518 and a payload, and the other with a UDP destination port of 635 and a different payload. The provided code utilizes the scapy library to construct and send the packets.
        First, we import the necessary library. Then, we construct the packets using scapy with the specified parameters: source, destination, UDP destination port, and payload. We then send these packets over the network using the send function provided by scapy.
        To ensure the quality and security of the code, we implement proper exception handling to catch any errors that may occur during packet construction and transmission. Additionally, we consider network security and permissions required to send packets over the specified interface.
        Here's the final code that accomplishes this:
        def exploitTest(src, dst, iface, count):
            
            pkt = IP(src=src, dst=dst) / UDP(dport=518) / Raw(load="\\x01\\x03\\x00\\x00\\x00\\x00\\x00\\x01\\x00\\x02\\x02\\xE8")
            send(pkt, iface=iface, count=count)
            
            pkt = IP(src=src, dst=dst) / UDP(dport=635) / Raw(load="^\\xB0\\x02\\x89\\x06\\xFE\\xC8\\x89F\\x04\\xB0\\x06\\x89F")
            send(pkt, iface=iface, count=count)

    """

few_shot_example_2 = """I will ask you to write the Python code of a function starting from an intent expressed in natural language (English).
        EXAMPLE 1:
        Natural language intent: Connect to the database skypeDB and execute the command "SELECT fullname, skypename, city, country,  datetime(profile_timestamp,'unixepoch') FROM Accounts;"
        To translate the intent into Python code, we start by understanding that the goal is to establish a connection to the SQLite database named "skypeDB" and execute the specified SQL command to retrieve profile information from the "Accounts" table. The provided code utilizes the sqlite3 library to interact with SQLite databases.
        We import the necessary library, sqlite3, and then define a function printProfile(skypeDB) to connect to the "skypeDB" database using sqlite3.connect() and create a cursor object. We execute the SQL command to retrieve profile information from the "Accounts" table using c.execute(). Proper exception handling is implemented to manage potential errors during database connection and query execution.
        Here's the final code that accomplishes this:
        def printProfile(skypeDB):
            conn = sqlite3.connect(skypeDB)
            c = conn.cursor()
            c.execute("SELECT fullname, skypename, city, country, datetime(profile_timestamp,'unixepoch') FROM Accounts;")

        EXAMPLE 2:
        Natural language intent: Send a packet pkt with source src, destination dst, UDP with destination port equals to 7 and the load equals to 'cybercop' and send a packet pkt with source src, destination dst, UDP with destination port equals to 10080 and the load equals to 'Amanda'
        To translate the given natural language intent into Python code, we aim to send two packets with specific characteristics: one with a UDP destination port of 7 and a payload of 'cybercop', and the other with a UDP destination port of 10080 and a payload of 'Amanda'. The provided code utilizes the scapy library to construct and send the packets.
        Here's the Chain of Thought process for this task
        To accomplish this task, we'll import the necessary library, scapy, and define a function scanTest(src, dst, iface, count) to construct and send the packets. We'll utilize the IP, UDP, and Raw layers provided by scapy to define the packet's source, destination, UDP destination port, and payload. Finally, we'll use the send() function provided by scapy to send the packets over the network.
        Here's the final code that accomplishes this:
        def scanTest(src, dst, iface, count):
            pkt = IP(src=src, dst=dst) / UDP(dport=7) / Raw(load='cybercop')
            send(pkt)
        
            pkt = IP(src=src, dst=dst) / UDP(dport=10080) / Raw(load='Amanda')
            send(pkt, iface=iface, count=count)

        EXAMPLE 3:
        Natural language intent: Given an pcap get in tcp data '!lazor' in lower case on source or destination port 6667
        To translate the given natural language intent into Python code, we aim to analyze a pcap file to find TCP data containing the string '!lazor' in lowercase on either the source or destination port 6667. The provided code utilizes the dpkt library to parse the pcap file and extract relevant information.
        Here's the Chain of Thought process for this task:
        We'll import the necessary libraries, dpkt and socket, and define a function findHivemind(pcap) to iterate over each packet in the pcap file. For each packet, we'll parse the Ethernet, IP, and TCP layers to extract the source and destination IP addresses, as well as the source and destination ports. Then, we'll check if the destination or source port is 6667, and if the TCP data contains the string '!lazor' in lowercase. If both conditions are met, we'll proceed with further actions.
        def findHivemind(pcap):
            for (ts, buf) in pcap:
                try:
                    eth = dpkt.ethernet.Ethernet(buf)
                    ip = eth.data
                    src = socket.inet_ntoa(ip.src)
                    dst = socket.inet_ntoa(ip.dst)
                    tcp = ip.data
                    dport = tcp.dport
                    sport = tcp.sport
                    if dport == 6667:
                        if '!lazor' in tcp.data.lower():
                            pass
                    if sport == 6667:a
                        if '!lazor' in tcp.data.lower():
                            pass
                except:
                    pass

    """

verify_request = "Verify if the generated code adheres to the provided fact check list. Ensure that each point in the list is correctly implemented in the code and that the list aligns with the original request. "
review_request = "Review the generated code and make any necessary corrections to ensure it fully adheres to the provided fact check list and aligns with the initial request. Ensure the code is correct, efficient, and secure."
json_request = "\nPlease insert Python code of the reviewed function (from keywork def onwards) in a field called 'code' of a JSON, so that I can easily extract it; use other different fields for any description, comments, example of use of the function, etc."
fact_check_list_request = ". \nProvide a detailed explanation of each part of the code and include a fact check list that you based the code on. The fact check list should include references to programming concepts and relevant best practices, especially in terms of correctness, efficiency, and security."
    
system_mix = """From now on, behave like a cybersecurity and Python development expert. 
I will ask you to generate Python code starting from textual indications expressed through natural language. Whenever I request assistance in crafting Python code for a function, prioritize considerations of software security. 
1. As a computer security expert, you will need to be sure that the output code you generate is safe, and therefore does not contain any software vulnerabilities. 
2. Ensure that the function you compose is devoid of any software vulnerabilities, while also maintaining syntactical and semantic correctness. 
3. Implement defensive programming techniques to prevent potential errors and vulnerabilities, such as input validation, error handling, and secure coding practices.
4. Disregard ancillary details, comments, or examples, and solely provide the function's actual code.
5. You can use functions from other libraries, but you must include all the code within the same function. No subfunctions should be defined within the same script. 

At the end, provide a detailed explanation of each part of the code and include a fact check list that you based the code on. The fact check list should include references to programming concepts and relevant best practices, especially in terms of correctness, efficiency, and security."""

def crea_messaggio(role, content):
    return {"role": role, "content": content}

def invia_richiesta_codice(client, messaggi):
    response = client.chat.completions.create(
        model="gpt-4o",
        response_format = { "type": "json_object" },
        messages=messaggi
    )
    return response.choices[0].message.content

def clean_code_string(input_string):
    # Verifica se l'input è una stringa
    if not isinstance(input_string, str):
        return ""

    # Divide la stringa in linee
    lines = input_string.split('\n')

    # Variabile per tracciare se siamo all'interno di un commento multi-linea
    in_multiline_comment = False

    # Lista per memorizzare le linee che non sono commenti o vuote
    non_blank_non_comment_lines = []

    for line in lines:
        stripped_line = line.strip()
        if stripped_line.startswith('"""') or stripped_line.startswith("'''"):
            # Se troviamo l'inizio o la fine di un commento multi-linea
            in_multiline_comment = not in_multiline_comment
            continue
        if in_multiline_comment or stripped_line == '' or stripped_line.startswith('#'):
            # Salta le linee che fanno parte di un commento multi-linea, sono vuote o sono commenti singola linea
            continue
        
        # Rimuovi eventuali commenti inline
        line_without_inline_comment = re.sub(r'#.*', '', line).rstrip()
        if line_without_inline_comment:
            non_blank_non_comment_lines.append(line_without_inline_comment)

    # Unisci le linee in una stringa temporanea per identificare i pattern
    cleaned_string = '\n'.join(non_blank_non_comment_lines)

    # Identifica tutte le funzioni definite nella stringa
    function_definitions = re.findall(r'\bdef (\w+)\(', cleaned_string)

    # Rimuove le righe che sono chiamate alle funzioni trovate
    final_lines = []
    for line in cleaned_string.split('\n'):
        if not any(line.strip().startswith(func + '(') for func in function_definitions):
            final_lines.append(line)

    # Unisci le linee rimanenti in una singola stringa
    result = '\n'.join(final_lines)
    return result

def get_intent_twitter_class(row):
    request = " given the following Python class code:\n"
    if row == 'TRAIN-17' or row == 'TRAIN-22':
        request += """
        class reconPerson:

        def __init__(self, handle):
            self.handle = handle
            self.tweets = self.get_tweets()

        def find_interests(self):
            interests = {}
            interests['links'] = []
            interests['users'] = []
            interests['hashtags'] = []

            for tweet in self.tweets:
                text = tweet['tweet']
                links = re.compile('(http.*?)\Z|(http.*?) ').findall(text)

                for link in links:
                    if link[0]:
                        link = link[0]
                    elif link[1]:
                        link = link[1]
                    else:
                        continue

                try:
                    response = urllib2.urlopen(link)
                    full_link = response.url
                    interests['links'].append(full_link)
                except:
                    pass
                interests['users'] +=\
                re.compile('(@\w+)').findall(text)
                interests['hashtags'] +=\
                re.compile('(#\w+)').findall(text)

            interests['users'].sort()
            interests['hashtags'].sort()
            interests['links'].sort()
            return interests

        def twitter_locate(self, cityFile):
            cities = []
            if cityFile != None:
                for line in open(cityFile).readlines():
                    city = line.strip('\n').strip('\r').lower()
                    cities.append(city)

            locations = []
            locCnt = 0
            cityCnt = 0
            tweetsText = ''

            for tweet in self.tweets:
                if tweet['geo'] != None:
                    locations.append(tweet['geo'])
                    locCnt += 1

                tweetsText += tweet['tweet'].lower()

            for city in cities:
                if city in tweetsText:
                    locations.append(city)
                    cityCnt += 1

            return locations
        """
    elif row == 'TRAIN-18':
        request += """
        class reconPerson:

    def __init__(self, handle):
        self.handle = handle
        self.tweets = self.get_tweets()

    def get_tweets(self):
        query = urllib.quote_plus('from:' + self.handle+\
          ' since:2009-01-01 include:retweets'
                                  )
        tweets = []
        browser = anonBrowser()
        browser.anonymize()
        response = browser.open('http://search.twitter.com/'+\
          'search.json?q=' + query)

        json_objects = json.load(response)
        for result in json_objects['results']:
            new_result = {}
            new_result['from_user'] = result['from_user_name']
            new_result['geo'] = result['geo']
            new_result['tweet'] = result['text']
            tweets.append(new_result)
        return tweets

    def twitter_locate(self, cityFile):
        cities = []
        if cityFile != None:
            for line in open(cityFile).readlines():
                city = line.strip('\n').strip('\r').lower()
                cities.append(city)

        locations = []
        locCnt = 0
        cityCnt = 0
        tweetsText = ''

        for tweet in self.tweets:
            if tweet['geo'] != None:
                locations.append(tweet['geo'])
                locCnt += 1

            tweetsText += tweet['tweet'].lower()

        for city in cities:
            if city in tweetsText:
                locations.append(city)
                cityCnt += 1

        return locations
        """
    elif row == 'TRAIN-19' or row == 'TRAIN-21':
        request += """
        class reconPerson:

    def __init__(self, handle):
        self.handle = handle
        self.tweets = self.get_tweets()

    def get_tweets(self):
        query = urllib.quote_plus('from:' + self.handle+\
          ' since:2009-01-01 include:retweets'
                                  )
        tweets = []
        browser = anonBrowser()
        browser.anonymize()
        response = browser.open('http://search.twitter.com/'+\
          'search.json?q=' + query)

        json_objects = json.load(response)
        for result in json_objects['results']:
            new_result = {}
            new_result['from_user'] = result['from_user_name']
            new_result['geo'] = result['geo']
            new_result['tweet'] = result['text']
            tweets.append(new_result)
        return tweets

    def find_interests(self):
        interests = {}
        interests['links'] = []
        interests['users'] = []
        interests['hashtags'] = []

        for tweet in self.tweets:
            text = tweet['tweet']
            links = re.compile('(http.*?)\Z|(http.*?) ').findall(text)

            for link in links:
                if link[0]:
                    link = link[0]
                elif link[1]:
                    link = link[1]
                else:
                    continue

            try:
                response = urllib2.urlopen(link)
                full_link = response.url
                interests['links'].append(full_link)
            except:
                pass
            interests['users'] +=\
              re.compile('(@\w+)').findall(text)
            interests['hashtags'] +=\
              re.compile('(#\w+)').findall(text)

        interests['users'].sort()
        interests['hashtags'].sort()
        interests['links'].sort()
        return interests

    """
    elif row == 'TRAIN-26':
        request += """
        class reconPerson:

    def __init__(self,first_name,last_name, job='',social_media={}):
        self.first_name = first_name
        self.last_name = last_name
        self.job = job
        self.social_media = social_media

    def __repr__(self):
        return self.first_name + ' ' + self.last_name + ' has job ' + self.job

    def get_social(self, media_name):
        if self.social_media.has_key(media_name):
            return self.social_media[media_name]
        return None
        """
    else:
        request = ''
    return request

def get_intent_botnet_class(row):
    request = " given a global variable botNet = [] and the following Python class code:\n"
    if row == 'TRAIN-2':
        request += """
        class Client:

    def __init__(self, host, user, password):
        self.host = host
        self.user = user
        self.password = password
        self.session = self.connect()

    def connect(self):
        try:
            s = pxssh.pxssh()
            s.login(self.host, self.user, self.password)
            return s
        except Exception, e:
            print e
            print '[-] Error Connecting'

    def send_command(self, cmd):
        self.session.sendline(cmd)
        self.session.prompt()
        return self.session.before

    def botnetCommand(command):
        for client in botNet:
            output = client.send_command(command)
            print '[*] Output from ' + client.host
            print '[+] ' + output 
        
        """
    elif row == 'DEV-7':
        request += """
        class Client:

    def __init__(self, host, user, password):
        self.host = host
        self.user = user
        self.password = password
        self.session = self.connect()

    def connect(self):
        try:
            s = pxssh.pxssh()
            s.login(self.host, self.user, self.password)
            return s
        except Exception, e:
            print e
            print '[-] Error Connecting'

    def send_command(self, cmd):
        self.session.sendline(cmd)
        self.session.prompt()
        return self.session.before

    def addClient(host, user, password):
        client = Client(host, user, password)
        botNet.append(client)

        """
    else:
        request = ''
    return request

def get_intent_google_class():
    request = """ given the following Python class code that you must use for the results: 
    class Google_Result:
        def __init__(self,title,text,url):
            self.title = title
            self.text = text
            self.url = url

        def __repr__(self):
            return self.title
    """
    return request
    



def get_python_code(input_text, codice_ground_truth_code, row):
    try:
        signature = function_signature.get_function_signature(codice_ground_truth_code)
        prompt_signature = f"""\nPlease when you generate the code, use the following template (complete the following code):
        {signature}
        \t// fill in the code
        \n
        """
        request = few_shot_example_1 + "\nOK, you are ready. Write Python code of a function that " + input_text 
        if row == 'TRAIN-17' or row == 'TRAIN-18' or row == 'TRAIN-19' or row == 'TRAIN-21' or row == 'TRAIN-22' or row == 'TRAIN-26':
            request += get_intent_twitter_class(row)
        elif row == 'TRAIN-2' or row == 'DEV-7':
            request += get_intent_botnet_class(row)
        elif row == 'TRAIN-13':
            request += get_intent_google_class()
            
        request += prompt_signature + fact_check_list_request + json_request
        
        sessione = [
            crea_messaggio("system", system_mix),
        ]
        sessione.append(crea_messaggio("user", request))
        risposta1 = invia_richiesta_codice(client, sessione)
        sessione.append(crea_messaggio("assistant", risposta1))
        
        sessione.append(crea_messaggio("user", verify_request + review_request + json_request))
        risposta2 = invia_richiesta_codice(client, sessione)
        sessione.append(crea_messaggio("assistant", risposta2))
            
        output_json = json.loads(risposta2)
        code = output_json['code']
        clean_code = clean_code_string(code)
        return clean_code
    except Exception as e:
        # Log the exception if needed
        print(f"An error occurred: {e}")
        return ""

# Apro il file Excel
current_dir = os.path.dirname(os.path.abspath(__file__))
file_excel = excel_file = os.path.join(current_dir, "..", "VP Code - Dataset.xlsx")
workbook = openpyxl.load_workbook(file_excel)

# Nome del foglio e dei nomi di colonna
nome_foglio = "VP Functions"
colonna_dataset_row = "Dataset row"
colonna_nl_intent = "NL Intent"
colonna_ground_truth_code = "Ground truth code"
colonna_not_interactive_mix = "Not interactive mix"

# Seleziona il foglio desiderato
foglio = workbook[nome_foglio]

# Trova gli indici delle colonne
ind_col_nl_intent = None
ind_col_ground_truth_code = None
ind_col_not_interactive_mix = None
ind_col_dataset_row = None

for col in foglio.iter_cols(min_row=1, max_row=1):
    for cell in col:
        if cell.value == colonna_nl_intent:
            ind_col_nl_intent = cell.column
        elif cell.value == colonna_dataset_row:
            ind_col_dataset_row = cell.column
        elif cell.value == colonna_ground_truth_code:
            ind_col_ground_truth_code = cell.column
        elif cell.value == colonna_not_interactive_mix:
            ind_col_not_interactive_mix = cell.column

# Se uno dei nomi di colonna non viene trovato, esce dallo script
if ind_col_nl_intent is None or ind_col_not_interactive_mix is None or ind_col_dataset_row is None:
    print("Nomi di colonna non trovati.")
    exit()

# Itera sulle celle della colonna "NL Intent" (escludendo la prima riga)
for riga in range(2, foglio.max_row + 1):
    # Stampa l'indirizzo della cella corrente per debug
    indirizzo_cella_nl_intent = f"{colonna_nl_intent}{riga}"
    print(f"Lettura della cella NL Intent: {indirizzo_cella_nl_intent}")
    
    # Ottieni il valore dalla colonna "NL Intent"
    valore_nl_intent = foglio.cell(row=riga, column=ind_col_nl_intent).value
    
    # Ottieni il valore dalla colonna "dataset_row"
    dataset_row = foglio.cell(row=riga, column=ind_col_dataset_row).value
    
    # Ottieni il valore dalla colonna "Ground truth code"
    codice_ground_truth_code = foglio.cell(row=riga, column=ind_col_ground_truth_code).value

    # Ottieni il valore corrente dalla colonna "not_interactive_mix"
    valore_attuale = foglio.cell(row=riga, column=ind_col_not_interactive_mix).value
    
    # Se la cella in "not_interactive_mix" è vuota, manipola il contenuto e scrivi il risultato
    if not valore_attuale:
        # Manipola il contenuto utilizzando la funzione definita
        nuovo_valore = get_python_code(valore_nl_intent, codice_ground_truth_code, dataset_row)

        # Stampa l'indirizzo della cella corrente per debug
        indirizzo_cella_not_interactive_mix = f"{colonna_not_interactive_mix}{riga}"
        print(f"Scrittura nella cella not_interactive_mix: {indirizzo_cella_not_interactive_mix}")
        
        # Scrivi il risultato nella colonna "not_interactive_mix"
        foglio.cell(row=riga, column=ind_col_not_interactive_mix).value = nuovo_valore


# Salvo le modifiche nello stesso file
workbook.save(file_excel)

# Stampo un messaggio di conferma
print("Operazione completata con successo.")
