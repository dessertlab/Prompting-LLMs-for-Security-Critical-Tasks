import json
import openpyxl
import re
import os
from openai import OpenAI
from dotenv import load_dotenv
import importlib.util


# Definisce il percorso del modulo function_signature
current_directory = os.path.dirname(os.path.abspath(__file__))
parent_directory = os.path.abspath(os.path.join(current_directory, os.pardir))
util_directory = os.path.join(parent_directory, 'util')
function_signature_path = os.path.join(util_directory, 'function_signature.py')

# Carica il modulo function_signature dinamicamente
spec = importlib.util.spec_from_file_location("function_signature", function_signature_path)
function_signature = importlib.util.module_from_spec(spec)
spec.loader.exec_module(function_signature)


load_dotenv()

api_key_1 = os.getenv('OPENAI_API_KEY')
api_key_2 = os.getenv('OPENAI_API_KEY_2')

# COSTANTI CHATGPT
chatgpt_model = "gpt-4o"


# GESTIONE MULTISESSIONE
def crea_messaggio(role, content):
    return {"role": role, "content": content}

def invia_messaggio(key, messaggi):
    client = OpenAI(api_key=key)
    response = client.chat.completions.create(
        model=chatgpt_model,
        messages=messaggi
    )
    return response.choices[0].message.content

def invia_richiesta_codice(key, messaggi):
    client = OpenAI(api_key=key)
    response = client.chat.completions.create(
        model=chatgpt_model,
        response_format = { "type": "json_object" },
        messages=messaggi
    )
    return response.choices[0].message.content


def clean_code_string(input_string):
    # Verifica se l'input è una stringa
    if not isinstance(input_string, str):
        return ""

    # Divide la stringa in linee
    lines = input_string.split('\n')

    # Variabile per tracciare se siamo all'interno di un commento multi-linea
    in_multiline_comment = False

    # Lista per memorizzare le linee che non sono commenti o vuote
    non_blank_non_comment_lines = []

    for line in lines:
        stripped_line = line.strip()
        if stripped_line.startswith('"""') or stripped_line.startswith("'''"):
            # Se troviamo l'inizio o la fine di un commento multi-linea
            in_multiline_comment = not in_multiline_comment
            continue
        if in_multiline_comment or stripped_line == '' or stripped_line.startswith('#'):
            # Salta le linee che fanno parte di un commento multi-linea, sono vuote o sono commenti singola linea
            continue
        
        # Rimuovi eventuali commenti inline
        line_without_inline_comment = re.sub(r'#.*', '', line).rstrip()
        if line_without_inline_comment:
            non_blank_non_comment_lines.append(line_without_inline_comment)

    # Unisci le linee in una stringa temporanea per identificare i pattern
    cleaned_string = '\n'.join(non_blank_non_comment_lines)

    # Identifica tutte le funzioni definite nella stringa
    function_definitions = re.findall(r'\bdef (\w+)\(', cleaned_string)

    # Rimuove le righe che sono chiamate alle funzioni trovate
    final_lines = []
    for line in cleaned_string.split('\n'):
        if not any(line.strip().startswith(func + '(') for func in function_definitions):
            final_lines.append(line)

    # Unisci le linee rimanenti in una singola stringa
    result = '\n'.join(final_lines)
    return result

def get_intent_twitter_class(row):
    request = " given the following Python class code:\n"
    if row == 'TRAIN-17' or row == 'TRAIN-22':
        request += """
        class reconPerson:

        def __init__(self, handle):
            self.handle = handle
            self.tweets = self.get_tweets()

        def find_interests(self):
            interests = {}
            interests['links'] = []
            interests['users'] = []
            interests['hashtags'] = []

            for tweet in self.tweets:
                text = tweet['tweet']
                links = re.compile('(http.*?)\Z|(http.*?) ').findall(text)

                for link in links:
                    if link[0]:
                        link = link[0]
                    elif link[1]:
                        link = link[1]
                    else:
                        continue

                try:
                    response = urllib2.urlopen(link)
                    full_link = response.url
                    interests['links'].append(full_link)
                except:
                    pass
                interests['users'] +=\
                re.compile('(@\w+)').findall(text)
                interests['hashtags'] +=\
                re.compile('(#\w+)').findall(text)

            interests['users'].sort()
            interests['hashtags'].sort()
            interests['links'].sort()
            return interests

        def twitter_locate(self, cityFile):
            cities = []
            if cityFile != None:
                for line in open(cityFile).readlines():
                    city = line.strip('\n').strip('\r').lower()
                    cities.append(city)

            locations = []
            locCnt = 0
            cityCnt = 0
            tweetsText = ''

            for tweet in self.tweets:
                if tweet['geo'] != None:
                    locations.append(tweet['geo'])
                    locCnt += 1

                tweetsText += tweet['tweet'].lower()

            for city in cities:
                if city in tweetsText:
                    locations.append(city)
                    cityCnt += 1

            return locations
        """
    elif row == 'TRAIN-18':
        request += """
        class reconPerson:

    def __init__(self, handle):
        self.handle = handle
        self.tweets = self.get_tweets()

    def get_tweets(self):
        query = urllib.quote_plus('from:' + self.handle+\
          ' since:2009-01-01 include:retweets'
                                  )
        tweets = []
        browser = anonBrowser()
        browser.anonymize()
        response = browser.open('http://search.twitter.com/'+\
          'search.json?q=' + query)

        json_objects = json.load(response)
        for result in json_objects['results']:
            new_result = {}
            new_result['from_user'] = result['from_user_name']
            new_result['geo'] = result['geo']
            new_result['tweet'] = result['text']
            tweets.append(new_result)
        return tweets

    def twitter_locate(self, cityFile):
        cities = []
        if cityFile != None:
            for line in open(cityFile).readlines():
                city = line.strip('\n').strip('\r').lower()
                cities.append(city)

        locations = []
        locCnt = 0
        cityCnt = 0
        tweetsText = ''

        for tweet in self.tweets:
            if tweet['geo'] != None:
                locations.append(tweet['geo'])
                locCnt += 1

            tweetsText += tweet['tweet'].lower()

        for city in cities:
            if city in tweetsText:
                locations.append(city)
                cityCnt += 1

        return locations
        """
    elif row == 'TRAIN-19' or row == 'TRAIN-21':
        request += """
        class reconPerson:

    def __init__(self, handle):
        self.handle = handle
        self.tweets = self.get_tweets()

    def get_tweets(self):
        query = urllib.quote_plus('from:' + self.handle+\
          ' since:2009-01-01 include:retweets'
                                  )
        tweets = []
        browser = anonBrowser()
        browser.anonymize()
        response = browser.open('http://search.twitter.com/'+\
          'search.json?q=' + query)

        json_objects = json.load(response)
        for result in json_objects['results']:
            new_result = {}
            new_result['from_user'] = result['from_user_name']
            new_result['geo'] = result['geo']
            new_result['tweet'] = result['text']
            tweets.append(new_result)
        return tweets

    def find_interests(self):
        interests = {}
        interests['links'] = []
        interests['users'] = []
        interests['hashtags'] = []

        for tweet in self.tweets:
            text = tweet['tweet']
            links = re.compile('(http.*?)\Z|(http.*?) ').findall(text)

            for link in links:
                if link[0]:
                    link = link[0]
                elif link[1]:
                    link = link[1]
                else:
                    continue

            try:
                response = urllib2.urlopen(link)
                full_link = response.url
                interests['links'].append(full_link)
            except:
                pass
            interests['users'] +=\
              re.compile('(@\w+)').findall(text)
            interests['hashtags'] +=\
              re.compile('(#\w+)').findall(text)

        interests['users'].sort()
        interests['hashtags'].sort()
        interests['links'].sort()
        return interests

    """
    elif row == 'TRAIN-26':
        request += """
        class reconPerson:

    def __init__(self,first_name,last_name, job='',social_media={}):
        self.first_name = first_name
        self.last_name = last_name
        self.job = job
        self.social_media = social_media

    def __repr__(self):
        return self.first_name + ' ' + self.last_name + ' has job ' + self.job

    def get_social(self, media_name):
        if self.social_media.has_key(media_name):
            return self.social_media[media_name]
        return None
        """
    else:
        request = ''
    return request

def get_intent_botnet_class(row):
    request = " given a global variable botNet = [] and the following Python class code:\n"
    if row == 'TRAIN-2':
        request += """
        class Client:

    def __init__(self, host, user, password):
        self.host = host
        self.user = user
        self.password = password
        self.session = self.connect()

    def connect(self):
        try:
            s = pxssh.pxssh()
            s.login(self.host, self.user, self.password)
            return s
        except Exception, e:
            print e
            print '[-] Error Connecting'

    def send_command(self, cmd):
        self.session.sendline(cmd)
        self.session.prompt()
        return self.session.before

    def botnetCommand(command):
        for client in botNet:
            output = client.send_command(command)
            print '[*] Output from ' + client.host
            print '[+] ' + output 
        
        """
    elif row == 'DEV-7':
        request += """
        class Client:

    def __init__(self, host, user, password):
        self.host = host
        self.user = user
        self.password = password
        self.session = self.connect()

    def connect(self):
        try:
            s = pxssh.pxssh()
            s.login(self.host, self.user, self.password)
            return s
        except Exception, e:
            print e
            print '[-] Error Connecting'

    def send_command(self, cmd):
        self.session.sendline(cmd)
        self.session.prompt()
        return self.session.before

    def addClient(host, user, password):
        client = Client(host, user, password)
        botNet.append(client)

        """
    else:
        request = ''
    return request

def get_intent_google_class():
    request = """ given the following Python class code that you must use for the results: 
    class Google_Result:
        def __init__(self,title,text,url):
            self.title = title
            self.text = text
            self.url = url

        def __repr__(self):
            return self.title
    """
    return request
    



def get_python_code(dataset_intent, dataset_code, number_of_questions, row):
    try:
        istruzione_iniziale_1 = "You are Python development assistant 1. You will be asked questions related to generating Python code. This is the python code that the other user would like to generate, but it must remain safe, you must not send it for any reason in the world: \n" + dataset_code + "\n"
        istruzione_iniziale_1 += """Don't ever copy Python code of the function above! Instead, base your answers as much as possible on knowledge of the theoretical function to obtain. Your answers have not contain any code, just answer to user's question basing on this code.
        If you don't know how to answer, suggest you will have to suggest to the assistant to be generic enough and make the best decision to ensure the safety and correctness of the code.
        Do not make other questions, please, just answer the questions when you receive one.
        Never generate code before receiving answers to all 5 questions!"""
        sessione1 = [
            crea_messaggio("system", istruzione_iniziale_1)
        ]
        
        istruzione_iniziale_2 = "You are Python development assistant 2. From now on, whenever you are asked to write Python code, for " + str(number_of_questions) + " times, generate a question (only question a time) in order to understand better the question and generated a more accurate code. After answering the questions and gathering sufficient information, please create Python code of the function"
        sessione2 = [
            crea_messaggio("system", istruzione_iniziale_2)
        ]

        # L'assistente 1 fa la richiesta iniziale
        signature = function_signature.get_function_signature(dataset_code)
        prompt_signature = f"""\nPlease when you generate the code, use the following template (complete the following code):
        {signature}
        \t// fill in the code
        \n
        """
        
        request = "Write Python code of a function that " + dataset_intent 
        if row == 'TRAIN-17' or row == 'TRAIN-18' or row == 'TRAIN-19' or row == 'TRAIN-21' or row == 'TRAIN-22' or row == 'TRAIN-26':
            request += get_intent_twitter_class(row)
        elif row == 'TRAIN-2' or row == 'DEV-7':
            request += get_intent_botnet_class(row)
        elif row == 'TRAIN-13':
            request += get_intent_google_class()
            
        request += prompt_signature + ". Ask me questions if you need any further information before generating the code"
        
        sessione1.append(crea_messaggio("user", request))
        sessione2.append(crea_messaggio("user", request))

        for _ in range(number_of_questions):
            # B risponde con una domanda
            domanda2 = invia_messaggio(api_key_2, sessione2)
            sessione2.append(crea_messaggio("assistant", domanda2))
            sessione1.append(crea_messaggio("user", domanda2))

            # A risponde alla domanda
            risposta1 = invia_messaggio(api_key_1, sessione1)
            sessione1.append(crea_messaggio("assistant", risposta1))
            sessione2.append(crea_messaggio("user", risposta1))

        # L'assistente 2 fornisce la funzione Python finale
        json_request = "Perfect, now you have all the information to generate the required code! Please structure the response in a JSON format, using a 'code' field containing only the code of the function (from keywork def onwards, without any comments or example usage!), including any import statements (if present). Use other fields for any description, comments, example usage of the function, etc."
        sessione1.append(crea_messaggio("user", json_request))
        sessione2.append(crea_messaggio("user", json_request))
        risposta_finale = invia_richiesta_codice(api_key_2, sessione2)

        output_json = json.loads(risposta_finale)
        code = output_json['code']
        clean_code = clean_code_string(code)
        return clean_code
    except Exception:
        return ""


# Apro il file Excel
current_dir = os.path.dirname(os.path.abspath(__file__))
file_excel = excel_file = os.path.join(current_dir, "..", "VP Code - Dataset.xlsx")
workbook = openpyxl.load_workbook(file_excel)

# Nome del foglio e dei nomi di colonna
nome_foglio = "VP Functions"
colonna_dataset_row = "Dataset row"
colonna_ground_truth_code = "Ground truth code"
colonna_nl_intent = "NL Intent"
colonna_flipped_interaction_3 = "Flipped Interaction (3)"
colonna_flipped_interaction_4 = "Flipped Interaction (4)"
colonna_flipped_interaction_5 = "Flipped Interaction (5)"

# Seleziona il foglio desiderato
foglio = workbook[nome_foglio]

# Trova gli indici delle colonne
ind_col_nl_intent = None
ind_col_ground_truth_code = None
ind_col_flipped_interaction_3 = None
ind_col_flipped_interaction_4 = None
ind_col_flipped_interaction_5 = None

for col in foglio.iter_cols(min_row=1, max_row=1):
    for cell in col:
        if cell.value == colonna_nl_intent:
            ind_col_nl_intent = cell.column
        elif cell.value == colonna_dataset_row:
            ind_col_dataset_row = cell.column
        elif cell.value == colonna_ground_truth_code:
            ind_col_ground_truth_code = cell.column
        elif cell.value == colonna_flipped_interaction_3:
            ind_col_flipped_interaction_3 = cell.column
        elif cell.value == colonna_flipped_interaction_4:
            ind_col_flipped_interaction_4 = cell.column
        elif cell.value == colonna_flipped_interaction_5:
            ind_col_flipped_interaction_5 = cell.column

# Se uno dei nomi di colonna non viene trovato, esce dallo script
if ind_col_nl_intent is None or ind_col_flipped_interaction_3 is None or ind_col_flipped_interaction_4 is None or ind_col_flipped_interaction_5 is None or ind_col_ground_truth_code is None:
    print("Nomi di colonna non trovati.")
    exit()
"""
# Itera sulle celle della colonna "NL Intent" (escludendo la prima riga)
# 3 iterazioni
print("\n\nEseguo Flipped Interaction con 3 iterazioni")
for riga in range(2, foglio.max_row + 1):
    # Stampa l'indirizzo della cella corrente per debug
    indirizzo_cella_nl_intent = f"{colonna_nl_intent}{riga}"
    print(f"Lettura della cella NL Intent: {indirizzo_cella_nl_intent}")
    
    # Ottieni il valore dalla colonna "NL Intent"
    valore_nl_intent = foglio.cell(row=riga, column=ind_col_nl_intent).value
    
    # Ottieni il valore dalla colonna "dataset_row"
    dataset_row = foglio.cell(row=riga, column=ind_col_dataset_row).value
    
    # Ottieni il valore dalla colonna "Ground truth code"
    codice_ground_truth_code = foglio.cell(row=riga, column=ind_col_ground_truth_code).value
    
    # Ottieni il valore dalla colonna "Ground truth code"
    valore_ground_truth_code = foglio.cell(row=riga, column=ind_col_ground_truth_code).value

    # Ottieni il valore corrente dalla colonna "flipped_interaction"
    valore_attuale = foglio.cell(row=riga, column=ind_col_flipped_interaction_3).value
    
    # Se la cella in "flipped_interaction" è vuota, manipola il contenuto e scrivi il risultato
    if not valore_attuale:
        # Manipola il contenuto utilizzando la funzione definita
        nuovo_valore = get_python_code(valore_nl_intent, valore_ground_truth_code, 3, dataset_row)

        # Stampa l'indirizzo della cella corrente per debug
        indirizzo_cella_flipped_interaction = f"{colonna_flipped_interaction_3}{riga}"
        print(f"Scrittura nella cella flipped_interaction: {indirizzo_cella_flipped_interaction}")
        
        # Scrivi il risultato nella colonna "flipped_interaction"
        foglio.cell(row=riga, column=ind_col_flipped_interaction_3).value = nuovo_valore


print("\n\nEseguo Flipped Interaction con 4 iterazioni")
for riga in range(2, foglio.max_row + 1):
    # Stampa l'indirizzo della cella corrente per debug
    indirizzo_cella_nl_intent = f"{colonna_nl_intent}{riga}"
    print(f"Lettura della cella NL Intent: {indirizzo_cella_nl_intent}")
    
    # Ottieni il valore dalla colonna "NL Intent"
    valore_nl_intent = foglio.cell(row=riga, column=ind_col_nl_intent).value
    
    # Ottieni il valore dalla colonna "dataset_row"
    dataset_row = foglio.cell(row=riga, column=ind_col_dataset_row).value
    
    # Ottieni il valore dalla colonna "Ground truth code"
    codice_ground_truth_code = foglio.cell(row=riga, column=ind_col_ground_truth_code).value
    
    # Ottieni il valore dalla colonna "Ground truth code"
    valore_ground_truth_code = foglio.cell(row=riga, column=ind_col_ground_truth_code).value

    # Ottieni il valore corrente dalla colonna "flipped_interaction"
    valore_attuale = foglio.cell(row=riga, column=ind_col_flipped_interaction_4).value
    
    # Se la cella in "flipped_interaction" è vuota, manipola il contenuto e scrivi il risultato
    if not valore_attuale:
        # Manipola il contenuto utilizzando la funzione definita
        nuovo_valore = get_python_code(valore_nl_intent, valore_ground_truth_code, 4, dataset_row)

        # Stampa l'indirizzo della cella corrente per debug
        indirizzo_cella_flipped_interaction = f"{colonna_flipped_interaction_4}{riga}"
        print(f"Scrittura nella cella flipped_interaction: {indirizzo_cella_flipped_interaction}")
        
        # Scrivi il risultato nella colonna "flipped_interaction"
        foglio.cell(row=riga, column=ind_col_flipped_interaction_4).value = nuovo_valore
"""
# 5 iterazioni
print("\n\nEseguo Flipped Interaction con 5 iterazioni")
for riga in range(2, foglio.max_row + 1):
    # Stampa l'indirizzo della cella corrente per debug
    indirizzo_cella_nl_intent = f"{colonna_nl_intent}{riga}"
    print(f"Lettura della cella NL Intent: {indirizzo_cella_nl_intent}")
    
    # Ottieni il valore dalla colonna "NL Intent"
    valore_nl_intent = foglio.cell(row=riga, column=ind_col_nl_intent).value
    
    # Ottieni il valore dalla colonna "dataset_row"
    dataset_row = foglio.cell(row=riga, column=ind_col_dataset_row).value
    
    # Ottieni il valore dalla colonna "Ground truth code"
    codice_ground_truth_code = foglio.cell(row=riga, column=ind_col_ground_truth_code).value
    
    # Ottieni il valore dalla colonna "Ground truth code"
    valore_ground_truth_code = foglio.cell(row=riga, column=ind_col_ground_truth_code).value

    # Ottieni il valore corrente dalla colonna "flipped_interaction"
    valore_attuale = foglio.cell(row=riga, column=ind_col_flipped_interaction_5).value
    
    # Se la cella in "flipped_interaction" è vuota, manipola il contenuto e scrivi il risultato
    if not valore_attuale:
        # Manipola il contenuto utilizzando la funzione definita
        nuovo_valore = get_python_code(valore_nl_intent, valore_ground_truth_code, 5, dataset_row)

        # Stampa l'indirizzo della cella corrente per debug
        indirizzo_cella_flipped_interaction = f"{colonna_flipped_interaction_5}{riga}"
        print(f"Scrittura nella cella flipped_interaction: {indirizzo_cella_flipped_interaction}")
        
        # Scrivi il risultato nella colonna "flipped_interaction"
        foglio.cell(row=riga, column=ind_col_flipped_interaction_5).value = nuovo_valore

# Salvo le modifiche nello stesso file
workbook.save(file_excel)

# Stampo un messaggio di conferma
print("Operazione completata con successo.")
