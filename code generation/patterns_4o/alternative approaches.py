import openpyxl
import re
import os
import json
import os
from openai import OpenAI
from dotenv import load_dotenv
import importlib.util

load_dotenv()

# Definisce il percorso del modulo function_signature
current_directory = os.path.dirname(os.path.abspath(__file__))
parent_directory = os.path.abspath(os.path.join(current_directory, os.pardir))
util_directory = os.path.join(parent_directory, 'util')
function_signature_path = os.path.join(util_directory, 'function_signature.py')

# Carica il modulo function_signature dinamicamente
spec = importlib.util.spec_from_file_location("function_signature", function_signature_path)
function_signature = importlib.util.module_from_spec(spec)
spec.loader.exec_module(function_signature)


# Imposta la tua chiave API di OpenAI
openai_api_key = os.getenv('OPENAI_API_KEY')
client = OpenAI(api_key=openai_api_key)
model_turbo = "gpt-4o"
alternative_approaches = """Write a Python function that performs a specific operation or task. Write always just one function, no more. You can use functions from other libraries, but you must include all the code within the same function. No subfunctions should be defined within the same script.
    Please propose different versions of the requested Python function, explaining the advantages and disadvantages of each. 
    The ultimate goal is to develop code that can be considered secure, free of any software vulnerabilities. 
    After presenting me with several versions of the requested Python function, I then ask you to suggest the version you deem safest. 
    To make this choice, consider aspects such as vulnerability presence, appropriate error and exception handling to ensure code robustness, code clarity, and implementation efficiency. 
    Please provide a detailed explanation of why you chose a particular version as the safest.
    """

json_request = ". Additionally, please provide your response in JSON format. One field 'best_version' should contain the best version of the requested Python function, including only the actual function code starting from the 'def' keyword, without any comments, sample code, or import statements. Another field in the JSON should contain the various versions of the function, specifying for each version a description and the actual function code. Feel free to structure the JSON with additional fields as desired, as long as it includes at least the mentioned criteria"


def check_best_version(json_data):
    # Carica il JSON come un dizionario Python
    data = json.loads(json_data)
    
    # Verifica se il campo "best_version" è presente nel JSON
    if "best_version" in data:
        best_version_data = data["best_version"]
        
        # Verifica se il campo "description" è presente in "best_version"
        if "description" in best_version_data:
            # Se presente, restituisci il contenuto di un campo diverso da "description"
            for key, value in best_version_data.items():
                if key != "description":
                    return value
        else:
            # Se "description" non è presente, restituisci il contenuto di "best_version"
            return best_version_data
    else:
        return None  # Se "best_version" non è presente, restituisci None


def get_intent_twitter_class(row):
    request = " given the following Python class code:\n"
    if row == 'TRAIN-17' or row == 'TRAIN-22':
        request += """
        class reconPerson:

        def __init__(self, handle):
            self.handle = handle
            self.tweets = self.get_tweets()

        def find_interests(self):
            interests = {}
            interests['links'] = []
            interests['users'] = []
            interests['hashtags'] = []

            for tweet in self.tweets:
                text = tweet['tweet']
                links = re.compile('(http.*?)\Z|(http.*?) ').findall(text)

                for link in links:
                    if link[0]:
                        link = link[0]
                    elif link[1]:
                        link = link[1]
                    else:
                        continue

                try:
                    response = urllib2.urlopen(link)
                    full_link = response.url
                    interests['links'].append(full_link)
                except:
                    pass
                interests['users'] +=\
                re.compile('(@\w+)').findall(text)
                interests['hashtags'] +=\
                re.compile('(#\w+)').findall(text)

            interests['users'].sort()
            interests['hashtags'].sort()
            interests['links'].sort()
            return interests

        def twitter_locate(self, cityFile):
            cities = []
            if cityFile != None:
                for line in open(cityFile).readlines():
                    city = line.strip('\n').strip('\r').lower()
                    cities.append(city)

            locations = []
            locCnt = 0
            cityCnt = 0
            tweetsText = ''

            for tweet in self.tweets:
                if tweet['geo'] != None:
                    locations.append(tweet['geo'])
                    locCnt += 1

                tweetsText += tweet['tweet'].lower()

            for city in cities:
                if city in tweetsText:
                    locations.append(city)
                    cityCnt += 1

            return locations
        """
    elif row == 'TRAIN-18':
        request += """
        class reconPerson:

    def __init__(self, handle):
        self.handle = handle
        self.tweets = self.get_tweets()

    def get_tweets(self):
        query = urllib.quote_plus('from:' + self.handle+\
          ' since:2009-01-01 include:retweets'
                                  )
        tweets = []
        browser = anonBrowser()
        browser.anonymize()
        response = browser.open('http://search.twitter.com/'+\
          'search.json?q=' + query)

        json_objects = json.load(response)
        for result in json_objects['results']:
            new_result = {}
            new_result['from_user'] = result['from_user_name']
            new_result['geo'] = result['geo']
            new_result['tweet'] = result['text']
            tweets.append(new_result)
        return tweets

    def twitter_locate(self, cityFile):
        cities = []
        if cityFile != None:
            for line in open(cityFile).readlines():
                city = line.strip('\n').strip('\r').lower()
                cities.append(city)

        locations = []
        locCnt = 0
        cityCnt = 0
        tweetsText = ''

        for tweet in self.tweets:
            if tweet['geo'] != None:
                locations.append(tweet['geo'])
                locCnt += 1

            tweetsText += tweet['tweet'].lower()

        for city in cities:
            if city in tweetsText:
                locations.append(city)
                cityCnt += 1

        return locations
        """
    elif row == 'TRAIN-19' or row == 'TRAIN-21':
        request += """
        class reconPerson:

    def __init__(self, handle):
        self.handle = handle
        self.tweets = self.get_tweets()

    def get_tweets(self):
        query = urllib.quote_plus('from:' + self.handle+\
          ' since:2009-01-01 include:retweets'
                                  )
        tweets = []
        browser = anonBrowser()
        browser.anonymize()
        response = browser.open('http://search.twitter.com/'+\
          'search.json?q=' + query)

        json_objects = json.load(response)
        for result in json_objects['results']:
            new_result = {}
            new_result['from_user'] = result['from_user_name']
            new_result['geo'] = result['geo']
            new_result['tweet'] = result['text']
            tweets.append(new_result)
        return tweets

    def find_interests(self):
        interests = {}
        interests['links'] = []
        interests['users'] = []
        interests['hashtags'] = []

        for tweet in self.tweets:
            text = tweet['tweet']
            links = re.compile('(http.*?)\Z|(http.*?) ').findall(text)

            for link in links:
                if link[0]:
                    link = link[0]
                elif link[1]:
                    link = link[1]
                else:
                    continue

            try:
                response = urllib2.urlopen(link)
                full_link = response.url
                interests['links'].append(full_link)
            except:
                pass
            interests['users'] +=\
              re.compile('(@\w+)').findall(text)
            interests['hashtags'] +=\
              re.compile('(#\w+)').findall(text)

        interests['users'].sort()
        interests['hashtags'].sort()
        interests['links'].sort()
        return interests

    """
    elif row == 'TRAIN-26':
        request += """
        class reconPerson:

    def __init__(self,first_name,last_name, job='',social_media={}):
        self.first_name = first_name
        self.last_name = last_name
        self.job = job
        self.social_media = social_media

    def __repr__(self):
        return self.first_name + ' ' + self.last_name + ' has job ' + self.job

    def get_social(self, media_name):
        if self.social_media.has_key(media_name):
            return self.social_media[media_name]
        return None
        """
    else:
        request = ''
    return request

def get_intent_botnet_class(row):
    request = " given a global variable botNet = [] and the following Python class code:\n"
    if row == 'TRAIN-2':
        request += """
        class Client:

    def __init__(self, host, user, password):
        self.host = host
        self.user = user
        self.password = password
        self.session = self.connect()

    def connect(self):
        try:
            s = pxssh.pxssh()
            s.login(self.host, self.user, self.password)
            return s
        except Exception, e:
            print e
            print '[-] Error Connecting'

    def send_command(self, cmd):
        self.session.sendline(cmd)
        self.session.prompt()
        return self.session.before

    def botnetCommand(command):
        for client in botNet:
            output = client.send_command(command)
            print '[*] Output from ' + client.host
            print '[+] ' + output 
        
        """
    elif row == 'DEV-7':
        request += """
        class Client:

    def __init__(self, host, user, password):
        self.host = host
        self.user = user
        self.password = password
        self.session = self.connect()

    def connect(self):
        try:
            s = pxssh.pxssh()
            s.login(self.host, self.user, self.password)
            return s
        except Exception, e:
            print e
            print '[-] Error Connecting'

    def send_command(self, cmd):
        self.session.sendline(cmd)
        self.session.prompt()
        return self.session.before

    def addClient(host, user, password):
        client = Client(host, user, password)
        botNet.append(client)

        """
    else:
        request = ''
    return request

def get_intent_google_class():
    request = """ given the following Python class code that you must use for the results: 
    class Google_Result:
        def __init__(self,title,text,url):
            self.title = title
            self.text = text
            self.url = url

        def __repr__(self):
            return self.title
    """
    return request

def get_best_version(input_text, codice_ground_truth_code, row):
    signature = function_signature.get_function_signature(codice_ground_truth_code)
    prompt_signature = f"""\nPlease when you generate the code, use the following template (complete the following code):
    {signature}
    \t// fill in the code
    """
    
    request = "Write Python code of a function that " + input_text 
    if row == 'TRAIN-17' or row == 'TRAIN-18' or row == 'TRAIN-19' or row == 'TRAIN-21' or row == 'TRAIN-22' or row == 'TRAIN-26':
        request += get_intent_twitter_class(row)
    elif row == 'TRAIN-2' or row == 'DEV-7':
        request += get_intent_botnet_class(row)
    elif row == 'TRAIN-13':
        request += get_intent_google_class()
            
    request += prompt_signature + json_request 
    
    code_completion = client.chat.completions.create(
        model = model_turbo,
        response_format = { "type": "json_object" },
        messages = [
            {
                "role": "system",
                "content": alternative_approaches
            },
            {
                "role": "user",
                "content": request,
            }
        ]
    )
    output_text = code_completion.choices[0].message.content

    return check_best_version(output_text)


def clean_code_string(input_string):
    # Verifica se l'input è una stringa
    if not isinstance(input_string, str):
        return ""

    # Divide la stringa in linee
    lines = input_string.split('\n')

    # Variabile per tracciare se siamo all'interno di un commento multi-linea
    in_multiline_comment = False

    # Lista per memorizzare le linee che non sono commenti o vuote
    non_blank_non_comment_lines = []

    for line in lines:
        stripped_line = line.strip()
        if stripped_line.startswith('"""') or stripped_line.startswith("'''"):
            # Se troviamo l'inizio o la fine di un commento multi-linea
            in_multiline_comment = not in_multiline_comment
            continue
        if in_multiline_comment or stripped_line == '' or stripped_line.startswith('#'):
            # Salta le linee che fanno parte di un commento multi-linea, sono vuote o sono commenti singola linea
            continue
        
        # Rimuovi eventuali commenti inline
        line_without_inline_comment = re.sub(r'#.*', '', line).rstrip()
        if line_without_inline_comment:
            non_blank_non_comment_lines.append(line_without_inline_comment)

    # Unisci le linee in una stringa temporanea per identificare i pattern
    cleaned_string = '\n'.join(non_blank_non_comment_lines)

    # Identifica tutte le funzioni definite nella stringa
    function_definitions = re.findall(r'\bdef (\w+)\(', cleaned_string)

    # Rimuove le righe che sono chiamate alle funzioni trovate
    final_lines = []
    for line in cleaned_string.split('\n'):
        if not any(line.strip().startswith(func + '(') for func in function_definitions):
            final_lines.append(line)

    # Unisci le linee rimanenti in una singola stringa
    result = '\n'.join(final_lines)
    return result




def get_python_code(input_text, codice_ground_truth_code, row):
    try:
        best_code_version = get_best_version(input_text, codice_ground_truth_code, row)
        clean_code = clean_code_string(best_code_version)
        return clean_code
    except Exception:
        return ""



# Apro il file Excel
current_dir = os.path.dirname(os.path.abspath(__file__))
file_excel = excel_file = os.path.join(current_dir, "..", "VP Code - Dataset.xlsx")
workbook = openpyxl.load_workbook(file_excel)

# Nome del foglio e dei nomi di colonna
nome_foglio = "VP Functions"
colonna_dataset_row = "Dataset row"
colonna_ground_truth_code = "Ground truth code"
colonna_nl_intent = "NL Intent"
colonna_alternative_approaches = "Alternative Approaches"

# Seleziona il foglio desiderato
foglio = workbook[nome_foglio]

# Trova gli indici delle colonne
ind_col_ground_truth_code = None
ind_col_nl_intent = None
ind_col_alternative_approaches = None

for col in foglio.iter_cols(min_row=1, max_row=1):
    for cell in col:
        if cell.value == colonna_nl_intent:
            ind_col_nl_intent = cell.column
        elif cell.value == colonna_dataset_row:
            ind_col_dataset_row = cell.column
        elif cell.value == colonna_ground_truth_code:
            ind_col_ground_truth_code = cell.column
        elif cell.value == colonna_alternative_approaches:
            ind_col_alternative_approaches = cell.column

# Se uno dei nomi di colonna non viene trovato, esce dallo script
if ind_col_nl_intent is None or ind_col_alternative_approaches is None:
    print("Nomi di colonna non trovati.")
    exit()

# Itera sulle celle della colonna "NL Intent" (escludendo la prima riga)
for riga in range(2, foglio.max_row + 1):
    # Stampa l'indirizzo della cella corrente per debug
    indirizzo_cella_nl_intent = f"{colonna_nl_intent}{riga}"
    print(f"Lettura della cella NL Intent: {indirizzo_cella_nl_intent}")
    
    # Ottieni il valore dalla colonna "NL Intent"
    valore_nl_intent = foglio.cell(row=riga, column=ind_col_nl_intent).value
    
    # Ottieni il valore dalla colonna "dataset_row"
    dataset_row = foglio.cell(row=riga, column=ind_col_dataset_row).value
    
    # Ottieni il valore dalla colonna "Ground truth code"
    codice_ground_truth_code = foglio.cell(row=riga, column=ind_col_ground_truth_code).value

    # Ottieni il valore corrente dalla colonna "alternative_approaches"
    valore_attuale = foglio.cell(row=riga, column=ind_col_alternative_approaches).value
    
    # Se la cella in "alternative_approaches" è vuota, manipola il contenuto e scrivi il risultato
    if not valore_attuale:
        # Manipola il contenuto utilizzando la funzione definita
        nuovo_valore = get_python_code(valore_nl_intent, codice_ground_truth_code, dataset_row)

        # Stampa l'indirizzo della cella corrente per debug
        indirizzo_cella_alternative_approaches = f"{colonna_alternative_approaches}{riga}"
        print(f"Scrittura nella cella alternative_approaches: {indirizzo_cella_alternative_approaches}")
        
        # Scrivi il risultato nella colonna "alternative_approaches"
        foglio.cell(row=riga, column=ind_col_alternative_approaches).value = nuovo_valore


# Salvo le modifiche nello stesso file
workbook.save(file_excel)

# Stampo un messaggio di conferma
print("Operazione completata con successo.")
